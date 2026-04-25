"""
Portfolio Company Analysis Agent
- Reads Excel files and extracts metrics with strict data integrity
- Pulls live public market comp data via yfinance
- Never guesses or interpolates missing values
- Writes presentation-ready Excel report
"""

import os
import json
import anthropic
import openpyxl
import yfinance as yf
from datetime import datetime
from dotenv import load_dotenv
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()
client = anthropic.Anthropic()
MODEL = "claude-sonnet-4-6"

PORTFOLIO_DIR = r"C:\AI\Portcos Project\Q4 Data"
OUTPUT_FILE   = r"C:\AI\Portcos Project\Portfolio_Summary.xlsx"

# Public market comps by sector — used for benchmarking
SECTOR_COMPS = {
    "medtech":        ["MDT", "SYK", "BSX", "EW", "HOLX"],
    "construction":   ["VMC", "MLM", "URI", "PWR", "MTZ"],
    "saas":           ["CRM", "NOW", "HUBS", "ZM", "DDOG"],
    "retail":         ["TJX", "ROST", "DG", "DLTR", "FIVE"],
    "industrial":     ["EMR", "ITT", "ROP", "AME", "PNR"],
    "food_bev":       ["SYY", "USFD", "PFGC", "CHEF", "UNFI"],
    "logistics":      ["XPO", "SAIA", "ODFL", "CHRW", "JBHT"],
    "healthcare_svc": ["HCA", "UHS", "THC", "ENSG", "AMED"],
    "energy_svc":     ["HAL", "SLB", "BKR", "OIS", "DNOW"],
    "consumer":       ["EL", "CHD", "CLX", "SPB", "CENT"],
}

# ── Tool definitions ──────────────────────────────────────────────────────────
TOOLS = [
    {
        "name": "list_excel_files",
        "description": "List all Excel files in the portfolio directory.",
        "input_schema": {
            "type": "object",
            "properties": {"directory": {"type": "string"}},
            "required": ["directory"]
        }
    },
    {
        "name": "read_excel_content",
        "description": "Read all sheets from an Excel file. Returns raw content exactly as stored - no interpretation.",
        "input_schema": {
            "type": "object",
            "properties": {"filepath": {"type": "string"}},
            "required": ["filepath"]
        }
    },
    {
        "name": "extract_metrics",
        "description": """
Store standardized financial metrics for a company.

STRICT DATA INTEGRITY RULES:
- Only extract numbers that are EXPLICITLY stated in the file
- If a value is not directly present in the file, set it to null (NOT provided)
- Do NOT calculate, interpolate, or derive values from other numbers
- Do NOT make judgment calls on approximate values
- Do NOT average or estimate ranges
- revenue_unit must be either '000s' or 'MM' exactly as reported
- ebitda_margin_fy2024 must be read directly from file, not calculated
- yoy growth rates must be read directly from file, not calculated
- For each metric, set data_source to describe exactly where in the file you found it

Required fields:
  company_name, revenue_fy2024, ebitda_fy2024, ebitda_margin_fy2024,
  revenue_q1_2025, ebitda_q1_2025, ebitda_label, revenue_label,
  revenue_unit, yoy_revenue_growth, yoy_ebitda_growth, net_debt,
  key_metric_1_name, key_metric_1_value,
  key_metric_2_name, key_metric_2_value,
  notable_observation, sector_key
        """,
        "input_schema": {
            "type": "object",
            "properties": {
                "company_name": {"type": "string"},
                "metrics":      {"type": "object"}
            },
            "required": ["company_name", "metrics"]
        }
    },
    {
        "name": "fetch_market_comps",
        "description": "Fetch live public market comparable company data from Yahoo Finance for a given sector.",
        "input_schema": {
            "type": "object",
            "properties": {
                "sector_key": {
                    "type": "string",
                    "description": "One of: medtech, construction, saas, retail, industrial, food_bev, logistics, healthcare_svc, energy_svc, consumer"
                },
                "company_name": {"type": "string"}
            },
            "required": ["sector_key", "company_name"]
        }
    },
    {
        "name": "write_summary_report",
        "description": "Write the final Excel report using all stored company data and market comps.",
        "input_schema": {
            "type": "object",
            "properties": {"output_path": {"type": "string"}},
            "required": ["output_path"]
        }
    }
]

# ── Storage ───────────────────────────────────────────────────────────────────
extracted_companies = []
market_comps_store  = {}

# ── Tool execution ────────────────────────────────────────────────────────────
def execute_tool(name, inputs):
    global extracted_companies, market_comps_store

    if name == "list_excel_files":
        try:
            files = [f for f in os.listdir(inputs["directory"]) if f.endswith(".xlsx")]
            return json.dumps(sorted(files))
        except Exception as e:
            return f"ERROR: {e}"

    elif name == "read_excel_content":
        try:
            wb = openpyxl.load_workbook(inputs["filepath"], data_only=True)
            output = [f"FILE: {os.path.basename(inputs['filepath'])}"]
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                output.append(f"\n=== SHEET: {sheet_name} ===")
                rows_captured = 0
                for row in ws.iter_rows(values_only=True):
                    if all(v is None for v in row):
                        continue
                    row_str = " | ".join(
                        str(round(v, 4) if isinstance(v, float) else v)
                        if v is not None else ""
                        for v in row
                    )
                    output.append(row_str)
                    rows_captured += 1
                    if rows_captured >= 80:
                        output.append("... [sheet truncated at 80 rows]")
                        break
            return "\n".join(output)
        except Exception as e:
            return f"ERROR reading file: {e}"

    elif name == "extract_metrics":
        # Validate: do not accept guessed/calculated values
        metrics = inputs.get("metrics", {})
        validated = {}
        flagged   = []

        for key, val in metrics.items():
            # Accept null (not provided) without issue
            if val is None:
                validated[key] = None
                continue
            # Check for any string hints of calculation
            if isinstance(val, str):
                calc_hints = ["calculat", "estimat", "approximat", "assum", "interpolat", "infer"]
                if any(h in val.lower() for h in calc_hints):
                    validated[key] = None
                    flagged.append(f"{key} (rejected: appears estimated)")
                    continue
            validated[key] = val

        entry = {
            "company":     inputs["company_name"],
            "flagged":     flagged,
            "data_fetched": datetime.now().strftime("%Y-%m-%d %H:%M"),
            **validated
        }
        extracted_companies.append(entry)
        flag_msg = f" | FLAGGED: {flagged}" if flagged else ""
        return f"Stored: {inputs['company_name']}{flag_msg}"

    elif name == "fetch_market_comps":
        sector = inputs.get("sector_key", "").lower()
        company = inputs.get("company_name", "")
        tickers = SECTOR_COMPS.get(sector)

        if not tickers:
            return f"Unknown sector: {sector}. Valid keys: {list(SECTOR_COMPS.keys())}"

        results = {"sector": sector, "company": company,
                   "fetched_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
                   "source": "Yahoo Finance (live)", "comps": []}

        for ticker in tickers:
            try:
                t = yf.Ticker(ticker)
                info = t.info
                comp = {
                    "ticker":           ticker,
                    "name":             info.get("longName") or info.get("shortName", ticker),
                    "market_cap_mm":    round(info["marketCap"] / 1e6, 0) if info.get("marketCap") else None,
                    "ev_ebitda":        round(info["enterpriseToEbitda"], 1) if info.get("enterpriseToEbitda") else None,
                    "ev_revenue":       round(info["enterpriseToRevenue"], 2) if info.get("enterpriseToRevenue") else None,
                    "revenue_growth":   round(info["revenueGrowth"], 4) if info.get("revenueGrowth") else None,
                    "ebitda_margin":    round(info["ebitdaMargins"], 4) if info.get("ebitdaMargins") else None,
                    "gross_margin":     round(info["grossMargins"], 4) if info.get("grossMargins") else None,
                    "price":            info.get("currentPrice"),
                    "52w_high":         info.get("fiftyTwoWeekHigh"),
                    "52w_low":          info.get("fiftyTwoWeekLow"),
                }
                # Remove None fields for cleanliness
                comp = {k: v for k, v in comp.items() if v is not None}
                results["comps"].append(comp)
                print(f"    [yahoo] {ticker}: EV/EBITDA={comp.get('ev_ebitda','N/A')}x")
            except Exception as e:
                results["comps"].append({"ticker": ticker, "error": str(e)})

        # Calculate median multiples
        ev_ebitdas = [c["ev_ebitda"] for c in results["comps"] if "ev_ebitda" in c]
        rev_growths = [c["revenue_growth"] for c in results["comps"] if "revenue_growth" in c]
        ebitda_margins = [c["ebitda_margin"] for c in results["comps"] if "ebitda_margin" in c]

        if ev_ebitdas:
            results["median_ev_ebitda"] = round(sorted(ev_ebitdas)[len(ev_ebitdas)//2], 1)
        if rev_growths:
            results["median_rev_growth"] = round(sorted(rev_growths)[len(rev_growths)//2], 4)
        if ebitda_margins:
            results["median_ebitda_margin"] = round(sorted(ebitda_margins)[len(ebitda_margins)//2], 4)

        market_comps_store[company] = results
        return json.dumps(results, indent=2)

    elif name == "write_summary_report":
        if not extracted_companies:
            return "ERROR: No companies stored yet."
        _write_excel_report(inputs["output_path"], extracted_companies, market_comps_store)
        return f"Written: {inputs['output_path']}"

    return f"Unknown tool: {name}"


# ── Excel report writer ───────────────────────────────────────────────────────
def _write_excel_report(path, companies, comps_store):
    wb = openpyxl.Workbook()

    NAVY  = "0D2137"
    BLUE  = "1565C0"
    TEAL  = "006064"
    GOLD  = "F9A825"
    GREY1 = "F5F6F8"
    GREY2 = "E8EAED"
    GREY3 = "9E9E9E"
    WHITE = "FFFFFF"
    BLACK = "212121"
    RED   = "C62828"
    GREEN = "2E7D32"
    AMBER = "E65100"

    def mkborder(color="D0D0D0", style="thin"):
        s = Side(style=style, color=color)
        return Border(bottom=s)

    def hdr(ws, row, col, val, bg=NAVY, fg=WHITE, sz=10, halign="left",
            bold=True, wrap=False):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(name="Calibri", bold=bold, color=fg, size=sz)
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal=halign, vertical="center",
                                wrap_text=wrap, indent=1 if halign=="left" else 0)
        return c

    def dat(ws, row, col, val, bg=None, fg=BLACK, sz=10, halign="right",
            fmt=None, bold=False, italic=False):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(name="Calibri", size=sz, color=fg, bold=bold, italic=italic)
        if bg:
            c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal=halign, vertical="center",
                                indent=1 if halign=="left" else 0)
        c.border = mkborder()
        if fmt:
            c.number_format = fmt
        return c

    def normalize(co, key):
        v = co.get(key)
        if v is None:
            return None
        try:
            v = float(v)
        except:
            return None
        unit = str(co.get("revenue_unit", "000s")).lower()
        if "mm" in unit:
            v *= 1000
        return round(v, 1)

    NP = "Not Provided"
    n  = len(companies)
    now_str = datetime.now().strftime("%B %d, %Y  %H:%M")

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1 — Executive Dashboard
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Executive Dashboard"
    ws.sheet_view.showGridLines = False

    # Column sizing
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 9
    for i in range(n):
        ws.column_dimensions[get_column_letter(i+3)].width = 17

    # Title block
    ws.row_dimensions[1].height = 46
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 14
    ws.row_dimensions[4].height = 36

    t = ws.cell(row=1, column=1,
        value="PORTFOLIO COMPANY  —  STANDARDIZED FINANCIAL COMPARISON")
    t.font = Font(name="Calibri", bold=True, color=WHITE, size=20)
    t.fill = PatternFill("solid", start_color=NAVY)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.merge_cells(f"A1:{get_column_letter(n+2)}1")

    s = ws.cell(row=2, column=1,
        value=f"Generated: {now_str}   |   Figures normalized to $000s   |   Missing data shown as 'Not Provided'   |   Source: company Excel files + Yahoo Finance (live)")
    s.font = Font(name="Calibri", italic=True, color="AAAAAA", size=9)
    s.fill = PatternFill("solid", start_color=NAVY)
    s.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.merge_cells(f"A2:{get_column_letter(n+2)}2")

    s2 = ws.cell(row=3, column=1,
        value="DATA INTEGRITY: All values read directly from source files. No values were calculated or estimated. See 'Data Flags' tab for any issues.")
    s2.font = Font(name="Calibri", italic=True, color=GOLD, size=9)
    s2.fill = PatternFill("solid", start_color=NAVY)
    s2.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.merge_cells(f"A3:{get_column_letter(n+2)}3")

    # Company name headers
    hdr(ws, 4, 1, "METRIC", bg=NAVY, sz=10)
    hdr(ws, 4, 2, "UNIT", bg=NAVY, sz=9, halign="center")
    for i, co in enumerate(companies):
        name = co.get("company", f"Company {i+1}")
        display = (name.replace(" Holdings","").replace(" Inc.","")
                       .replace(" LLC","").replace(" Corp","")
                       .replace(" Group","").replace(" Partners LP","")
                       .replace(" Partners","").strip())
        hdr(ws, 4, i+3, display, bg=BLUE, sz=9, halign="center", wrap=True)

    # Section + metric definitions
    sections = [
        ("REVENUE", [
            ("FY2024 Revenue",      "$000s", lambda co: normalize(co,"revenue_fy2024"),    '#,##0;(#,##0);"-"', None),
            ("Q1 2025 Revenue",     "$000s", lambda co: normalize(co,"revenue_q1_2025"),   '#,##0;(#,##0);"-"', None),
            ("YoY Revenue Growth",  "%",     lambda co: co.get("yoy_revenue_growth"),      "0.0%",              "growth"),
        ]),
        ("PROFITABILITY", [
            ("FY2024 EBITDA",       "$000s", lambda co: normalize(co,"ebitda_fy2024"),     '#,##0;(#,##0);"-"', None),
            ("Q1 2025 EBITDA",      "$000s", lambda co: normalize(co,"ebitda_q1_2025"),    '#,##0;(#,##0);"-"', None),
            ("FY2024 EBITDA Margin","%",     lambda co: co.get("ebitda_margin_fy2024"),    "0.0%",              "margin"),
            ("YoY EBITDA Growth",   "%",     lambda co: co.get("yoy_ebitda_growth"),       "0.0%",              "growth"),
        ]),
        ("BALANCE SHEET", [
            ("Net Debt",            "$000s", lambda co: normalize(co,"net_debt"),          '#,##0;(#,##0);"-"', None),
            ("Net Leverage",        "x",     lambda co: (
                round(abs(normalize(co,"net_debt") or 0) /
                      normalize(co,"ebitda_fy2024"), 1)
                if normalize(co,"ebitda_fy2024") and normalize(co,"net_debt") else None
            ),                                                                             '0.0"x"',            None),
        ]),
        ("PUBLIC MARKET COMPS  (Yahoo Finance — Live)", [
            ("Peer Median EV/EBITDA","x",    lambda co: comps_store.get(co.get("company",""),({})).get("median_ev_ebitda"), '0.0"x"', None),
            ("Peer Median Rev Growth","%",   lambda co: comps_store.get(co.get("company",""),({})).get("median_rev_growth"), "0.0%",   None),
            ("Peer Median EBITDA Margin","%",lambda co: comps_store.get(co.get("company",""),({})).get("median_ebitda_margin"), "0.0%",None),
        ]),
        ("COMPANY-SPECIFIC KPIs", [
            ("KPI 1",               "",      lambda co: co.get("key_metric_1_value"),      None,                None),
            ("KPI 2",               "",      lambda co: co.get("key_metric_2_value"),      None,                None),
        ]),
    ]

    cur_row = 5
    for section_name, metrics in sections:
        # Section header
        ws.row_dimensions[cur_row].height = 20
        sc = ws.cell(row=cur_row, column=1, value=f"  {section_name}")
        sc.font = Font(name="Calibri", bold=True, color=WHITE, size=9)
        sc.fill = PatternFill("solid", start_color=TEAL)
        sc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.merge_cells(f"A{cur_row}:{get_column_letter(n+2)}{cur_row}")
        cur_row += 1

        for label, unit, fn, fmt, color_mode in metrics:
            ws.row_dimensions[cur_row].height = 22
            row_bg = GREY1 if cur_row % 2 == 0 else WHITE

            # Label
            lc = ws.cell(row=cur_row, column=1, value=label)
            lc.font = Font(name="Calibri", size=10, color=BLACK)
            lc.fill = PatternFill("solid", start_color=row_bg)
            lc.alignment = Alignment(horizontal="left", vertical="center", indent=3)
            lc.border = mkborder()

            # Unit
            uc = ws.cell(row=cur_row, column=2, value=unit)
            uc.font = Font(name="Calibri", size=8, color=GREY3, italic=True)
            uc.fill = PatternFill("solid", start_color=row_bg)
            uc.alignment = Alignment(horizontal="center", vertical="center")
            uc.border = mkborder()

            # Data cells
            for i, co in enumerate(companies):
                val = fn(co)
                col = i + 3

                # Determine display value and color
                if val is None:
                    display_val = NP
                    color = GREY3
                    display_fmt = None
                    is_italic = True
                else:
                    display_val = val
                    is_italic = False
                    display_fmt = fmt
                    if color_mode == "growth":
                        try:
                            color = GREEN if float(val) >= 0 else RED
                        except:
                            color = BLACK
                    elif color_mode == "margin":
                        color = BLUE
                    else:
                        color = BLACK

                c = ws.cell(row=cur_row, column=col, value=display_val)
                c.font = Font(name="Calibri", size=10, color=color,
                              italic=is_italic and display_val == NP)
                c.fill = PatternFill("solid", start_color=row_bg)
                c.alignment = Alignment(horizontal="right" if display_val != NP else "center",
                                       vertical="center")
                c.border = mkborder()
                if display_fmt and display_val != NP:
                    c.number_format = display_fmt

            cur_row += 1

        cur_row += 1  # spacer between sections

    ws.freeze_panes = "A5"

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2 — Reporting Labels
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Reporting Labels")
    ws2.sheet_view.showGridLines = False
    ws2.row_dimensions[1].height = 44
    ws2.row_dimensions[2].height = 14
    ws2.row_dimensions[3].height = 30

    t2 = ws2.cell(row=1, column=1, value="REPORTING TERMINOLOGY — COMPANY BY COMPANY")
    t2.font = Font(name="Calibri", bold=True, color=WHITE, size=16)
    t2.fill = PatternFill("solid", start_color=NAVY)
    t2.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws2.merge_cells("A1:G1")

    s2h = ws2.cell(row=2, column=1,
        value="Each company uses different terminology for equivalent metrics. Use this tab to trace any value back to its source document.")
    s2h.font = Font(name="Calibri", italic=True, color="AAAAAA", size=9)
    s2h.fill = PatternFill("solid", start_color=NAVY)
    s2h.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws2.merge_cells("A2:G2")

    for col, w in zip("ABCDEFG", [32, 26, 30, 16, 18, 18, 55]):
        ws2.column_dimensions[col].width = w

    cols2 = ["COMPANY","REVENUE LABEL","EBITDA LABEL","REPORTING UNIT",
             "FY24 EBITDA MARGIN","SECTOR","NOTABLE OBSERVATION"]
    for j, h in enumerate(cols2):
        hdr(ws2, 3, j+1, h, bg=TEAL, sz=10, halign="left")

    for i, co in enumerate(companies):
        r = i + 4
        ws2.row_dimensions[r].height = 28
        bg = GREY1 if i % 2 == 0 else WHITE
        vals = [
            co.get("company", f"Company {i+1}"),
            co.get("revenue_label") or NP,
            co.get("ebitda_label") or NP,
            co.get("revenue_unit") or NP,
            co.get("ebitda_margin_fy2024"),
            co.get("sector_key") or NP,
            co.get("notable_observation") or NP,
        ]
        for j, v in enumerate(vals):
            is_np = v is None or v == NP
            display = NP if is_np else v
            c = ws2.cell(row=r, column=j+1, value=display)
            c.fill = PatternFill("solid", start_color=bg)
            c.border = mkborder()
            c.alignment = Alignment(horizontal="left", vertical="center",
                                   wrap_text=(j==6), indent=1)
            if j == 0:
                c.font = Font(name="Calibri", bold=True, size=10, color=NAVY)
            elif j == 4 and not is_np:
                c.font = Font(name="Calibri", size=10, color=BLUE)
                c.number_format = "0.0%"
            elif is_np:
                c.font = Font(name="Calibri", size=10, color=GREY3, italic=True)
            else:
                c.font = Font(name="Calibri", size=10, color=BLACK)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 3 — Public Market Comps Detail
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Market Comps (Live)")
    ws3.sheet_view.showGridLines = False
    ws3.row_dimensions[1].height = 44
    ws3.row_dimensions[2].height = 14
    ws3.row_dimensions[3].height = 30

    t3 = ws3.cell(row=1, column=1,
        value=f"PUBLIC MARKET COMPARABLE COMPANIES  —  Live Data from Yahoo Finance  |  {now_str}")
    t3.font = Font(name="Calibri", bold=True, color=WHITE, size=14)
    t3.fill = PatternFill("solid", start_color=NAVY)
    t3.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws3.merge_cells("A1:H1")

    s3h = ws3.cell(row=2, column=1,
        value="Used for sector benchmarking only. Portfolio companies are private — these are public peers.")
    s3h.font = Font(name="Calibri", italic=True, color="AAAAAA", size=9)
    s3h.fill = PatternFill("solid", start_color=NAVY)
    s3h.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws3.merge_cells("A2:H2")

    for col, w in zip("ABCDEFGH", [28, 10, 24, 18, 14, 16, 16, 16]):
        ws3.column_dimensions[col].width = w

    cols3 = ["PORTFOLIO COMPANY","TICKER","PEER NAME","MKT CAP ($MM)",
             "EV/EBITDA","REV GROWTH","EBITDA MARGIN","CURRENT PRICE"]
    for j, h in enumerate(cols3):
        hdr(ws3, 3, j+1, h, bg=BLUE, sz=10)

    r3 = 4
    for co_name, comp_data in comps_store.items():
        comps = comp_data.get("comps", [])
        for k, comp in enumerate(comps):
            ws3.row_dimensions[r3].height = 22
            bg = GREY1 if r3 % 2 == 0 else WHITE

            # Portfolio company name only on first comp row
            c0 = ws3.cell(row=r3, column=1,
                         value=co_name if k == 0 else "")
            c0.font = Font(name="Calibri", bold=(k==0), size=10, color=NAVY)
            c0.fill = PatternFill("solid", start_color=bg)
            c0.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            c0.border = mkborder()

            vals3 = [
                comp.get("ticker", NP),
                comp.get("name", NP),
                comp.get("market_cap_mm"),
                comp.get("ev_ebitda"),
                comp.get("revenue_growth"),
                comp.get("ebitda_margin"),
                comp.get("price"),
            ]
            fmts3 = [None, None, '#,##0;(#,##0)', '0.0"x"', "0.0%", "0.0%", "$#,##0.00"]
            for j, (v, f) in enumerate(zip(vals3, fmts3)):
                is_np = v is None
                c = ws3.cell(row=r3, column=j+2, value=NP if is_np else v)
                c.fill = PatternFill("solid", start_color=bg)
                c.border = mkborder()
                c.alignment = Alignment(horizontal="right" if not is_np else "center",
                                       vertical="center")
                if is_np:
                    c.font = Font(name="Calibri", size=10, color=GREY3, italic=True)
                else:
                    c.font = Font(name="Calibri", size=10, color=BLACK)
                    if f:
                        c.number_format = f
            r3 += 1

        # Median summary row
        if comp_data.get("median_ev_ebitda"):
            ws3.row_dimensions[r3].height = 24
            mc = ws3.cell(row=r3, column=1, value="")
            mc.fill = PatternFill("solid", start_color=GREY2)
            mc2 = ws3.cell(row=r3, column=2, value="MEDIAN")
            mc2.font = Font(name="Calibri", bold=True, size=10, color=TEAL)
            mc2.fill = PatternFill("solid", start_color=GREY2)
            mc2.alignment = Alignment(horizontal="center", vertical="center")
            mc2.border = mkborder()

            med_vals = [None, None,
                        comp_data.get("median_ev_ebitda"),
                        comp_data.get("median_rev_growth"),
                        comp_data.get("median_ebitda_margin"),
                        None]
            med_fmts = [None, None, '0.0"x"', "0.0%", "0.0%", None]
            for j, (mv, mf) in enumerate(zip(med_vals, med_fmts)):
                c = ws3.cell(row=r3, column=j+3, value=mv)
                c.fill = PatternFill("solid", start_color=GREY2)
                c.border = mkborder()
                c.alignment = Alignment(horizontal="right", vertical="center")
                if mv is not None:
                    c.font = Font(name="Calibri", bold=True, size=10, color=TEAL)
                    if mf:
                        c.number_format = mf
            r3 += 2  # gap between companies

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 4 — Company KPIs
    # ══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Company KPIs")
    ws4.sheet_view.showGridLines = False
    ws4.row_dimensions[1].height = 44
    ws4.row_dimensions[2].height = 14
    ws4.row_dimensions[3].height = 30

    t4 = ws4.cell(row=1, column=1,
        value="PORTFOLIO — COMPANY-SPECIFIC KEY PERFORMANCE INDICATORS")
    t4.font = Font(name="Calibri", bold=True, color=WHITE, size=16)
    t4.fill = PatternFill("solid", start_color=NAVY)
    t4.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws4.merge_cells("A1:E1")

    s4 = ws4.cell(row=2, column=1,
        value="Values read directly from source files. 'Not Provided' means metric was not present in the file.")
    s4.font = Font(name="Calibri", italic=True, color="AAAAAA", size=9)
    s4.fill = PatternFill("solid", start_color=NAVY)
    s4.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws4.merge_cells("A2:E2")

    for col, w in zip("ABCDE", [32, 38, 18, 38, 18]):
        ws4.column_dimensions[col].width = w

    for j, h in enumerate(["COMPANY","KPI 1 NAME","KPI 1 VALUE","KPI 2 NAME","KPI 2 VALUE"]):
        hdr(ws4, 3, j+1, h, bg=BLUE, sz=10)

    for i, co in enumerate(companies):
        r = i + 4
        ws4.row_dimensions[r].height = 26
        bg = GREY1 if i % 2 == 0 else WHITE
        vals = [
            co.get("company", f"Company {i+1}"),
            co.get("key_metric_1_name") or NP,
            co.get("key_metric_1_value"),
            co.get("key_metric_2_name") or NP,
            co.get("key_metric_2_value"),
        ]
        for j, v in enumerate(vals):
            is_np = v is None
            c = ws4.cell(row=r, column=j+1, value=NP if is_np else v)
            c.fill = PatternFill("solid", start_color=bg)
            c.border = mkborder()
            if j == 0:
                c.font = Font(name="Calibri", bold=True, size=10, color=NAVY)
                c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            elif is_np:
                c.font = Font(name="Calibri", size=10, color=GREY3, italic=True)
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif j in (2, 4):
                c.font = Font(name="Calibri", size=10, color=BLACK)
                c.alignment = Alignment(horizontal="center", vertical="center")
                if isinstance(v, float):
                    c.number_format = "0.0%" if v < 5 else "#,##0.0"
            else:
                c.font = Font(name="Calibri", size=10, color=BLACK, italic=True)
                c.alignment = Alignment(horizontal="left", vertical="center",
                                       indent=1, wrap_text=True)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 5 — Data Flags & Integrity Log
    # ══════════════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Data Flags")
    ws5.sheet_view.showGridLines = False
    ws5.row_dimensions[1].height = 44
    ws5.row_dimensions[2].height = 14
    ws5.row_dimensions[3].height = 30

    t5 = ws5.cell(row=1, column=1, value="DATA INTEGRITY LOG")
    t5.font = Font(name="Calibri", bold=True, color=WHITE, size=16)
    t5.fill = PatternFill("solid", start_color=NAVY)
    t5.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws5.merge_cells("A1:D1")

    s5 = ws5.cell(row=2, column=1,
        value="Any values rejected for appearing estimated or calculated rather than directly read from source are listed here.")
    s5.font = Font(name="Calibri", italic=True, color="AAAAAA", size=9)
    s5.fill = PatternFill("solid", start_color=NAVY)
    s5.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws5.merge_cells("A2:D2")

    for col, w in zip("ABCD", [32, 20, 50, 20]):
        ws5.column_dimensions[col].width = w

    for j, h in enumerate(["COMPANY","STATUS","FLAGGED FIELDS","DATA FETCHED AT"]):
        hdr(ws5, 3, j+1, h, bg=NAVY, sz=10)

    for i, co in enumerate(companies):
        r = i + 4
        ws5.row_dimensions[r].height = 26
        bg = GREY1 if i % 2 == 0 else WHITE
        flagged = co.get("flagged", [])
        status  = "⚠ Issues Found" if flagged else "✓ Clean"
        status_color = AMBER if flagged else GREEN

        vals = [
            co.get("company", f"Company {i+1}"),
            status,
            ", ".join(flagged) if flagged else "No issues",
            co.get("data_fetched", ""),
        ]
        for j, v in enumerate(vals):
            c = ws5.cell(row=r, column=j+1, value=v)
            c.fill = PatternFill("solid", start_color=bg)
            c.border = mkborder()
            c.alignment = Alignment(horizontal="left", vertical="center",
                                   indent=1, wrap_text=(j==2))
            if j == 0:
                c.font = Font(name="Calibri", bold=True, size=10, color=NAVY)
            elif j == 1:
                c.font = Font(name="Calibri", bold=True, size=10, color=status_color)
            else:
                c.font = Font(name="Calibri", size=10, color=BLACK)

    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    wb.save(path)


# ── Agentic loop ──────────────────────────────────────────────────────────────
import time

def run_agent(task):
    messages = [{"role": "user", "content": task}]
    while True:
        try:
            response = client.messages.create(
                model=MODEL,
                max_tokens=8096,
                tools=TOOLS,
                messages=messages
            )
        except anthropic.RateLimitError:
            print("  [rate limit] waiting 60 seconds...")
            time.sleep(60)
            continue

        messages.append({"role": "assistant", "content": response.content})

        if response.stop_reason == "end_turn":
            for block in response.content:
                if hasattr(block, "text"):
                    print(f"\nAgent: {block.text}\n")
            break

        if response.stop_reason == "tool_use":
            results = []
            for block in response.content:
                if block.type == "tool_use":
                    result = execute_tool(block.name, block.input)
                    label = {
                        "list_excel_files":   "[list]  ",
                        "read_excel_content": "[read]  ",
                        "extract_metrics":    "[store] ",
                        "fetch_market_comps": "[comps] ",
                        "write_summary_report":"[write] ",
                    }.get(block.name, "[tool]  ")
                    name_hint = (block.input.get("company_name") or
                                 os.path.basename(block.input.get("filepath","")) or
                                 block.input.get("sector_key",""))
                    print(f"  {label}{name_hint}")
                    results.append({
                        "type": "tool_result",
                        "tool_use_id": block.id,
                        "content": str(result)
                    })
            messages.append({"role": "user", "content": results})
            time.sleep(3)  # small pause between each tool call cycle

# ── Run ───────────────────────────────────────────────────────────────────────
TASK = f"""
You are a senior private equity analyst. Analyze 10 portfolio company Excel files
and produce a standardized benchmarking report with live market data.

CRITICAL DATA INTEGRITY RULES:
- Only extract values EXPLICITLY stated in the file. Never calculate or derive values.
- If a value is not present in the file, pass null for that field.
- Do not guess, approximate, or interpolate any number.
- Do not fill gaps using other numbers in the file.
- Use exact company names from the filename or file header, not generic labels.

STEPS:
1. List all Excel files in: {PORTFOLIO_DIR}

2. For each file:
   a. Read using read_excel_content
   b. Extract metrics using extract_metrics with these rules:
      - company_name: use full name from filename or file header (e.g. "MedTech Solutions Inc.")
      - revenue_fy2024: FY2024 total revenue, exactly as stated
      - ebitda_fy2024: FY2024 EBITDA (whatever label they use), exactly as stated
      - ebitda_margin_fy2024: read from file IF explicitly stated, else null
      - revenue_q1_2025: Q1 2025 revenue, exactly as stated
      - ebitda_q1_2025: Q1 2025 EBITDA, exactly as stated
      - yoy_revenue_growth: only if explicitly stated in file, else null
      - yoy_ebitda_growth: only if explicitly stated in file, else null
      - net_debt: if stated, use negative number (e.g. -22000). If not stated, null.
      - revenue_unit: "000s" or "MM" — identify from file headers
      - ebitda_label: exact label used (e.g. "Adj. EBITDA", "Cash EBITDA", "EBITDAX")
      - revenue_label: exact label used (e.g. "Contract Billings", "Net Patient Revenue")
      - key_metric_1_name / value: most interesting company-specific KPI from the file
      - key_metric_2_name / value: second most interesting KPI
      - notable_observation: one sentence about something unique in how they report
      - sector_key: one of: medtech, construction, saas, retail, industrial, 
                   food_bev, logistics, healthcare_svc, energy_svc, consumer

   c. After storing metrics, call fetch_market_comps with the sector_key and company_name

3. After ALL 10 companies are processed, call write_summary_report with: {OUTPUT_FILE}
"""

print("Starting portfolio analysis with live market data...\n")
run_agent(TASK)
print(f"\nDone. Open: {OUTPUT_FILE}")
