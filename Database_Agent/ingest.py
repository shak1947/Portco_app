"""
ingest.py — Deterministic parser for PE portfolio Excel files.

Ingests all 10 portfolio company Excel files from Portco data/ folder.
No AI mapping needed — structure is deterministic.
"""

import os
import re
from typing import Optional, Tuple, List, Dict
import openpyxl
from sqlalchemy import text as sqla_text


# ================================================================================
# CANONICAL MAPPINGS
# ================================================================================

CANONICAL_METRIC_MAP = {
    # Revenue variants → "revenue"
    "net revenue": "revenue",
    "contract billings": "revenue",
    "total revenue": "revenue",
    "net sales": "revenue",
    "total net revenue": "revenue",
    "net patient revenue": "revenue",
    "total revenue ($000s)": "revenue",
    "total revenue ($mm)": "revenue",

    # EBITDA variants (priority order in find_metric_row)
    "adjusted ebitda": "ebitda",
    "adj. ebitda": "ebitda",
    "cash ebitda": "ebitda",
    "brand ebitda": "ebitda",
    "ebitdax": "ebitdax",
    "run-rate ebitda": "ebitda_run_rate",
    "reported ebitda": "ebitda_reported",
    "ebitda": "ebitda",
    "four-wall ebitda": "ebitda_four_wall",

    # Other useful rows
    "gross profit": "gross_profit",
    "gross margin $": "gross_profit",
}

DCF_ASSUMPTION_MAP = {
    "revenue growth (yr 1-3)": "rev_growth_yr1_3",
    "revenue growth (yr 4-5)": "rev_growth_yr4_5",
    "revenue growth yr1-2": "rev_growth_yr1_2",
    "revenue growth yr3-5": "rev_growth_yr3_5",
    "billings growth (yr 1-3)": "rev_growth_yr1_3",
    "billings growth (yr 4-5)": "rev_growth_yr4_5",
    "arr growth yr 1-2": "rev_growth_yr1_2",
    "arr growth yr 3-4": "rev_growth_yr3_4",
    "arr growth yr 5": "rev_growth_yr5",

    "ebitda margin (stabilized)": "ebitda_margin_terminal",
    "ebitda margin": "ebitda_margin_terminal",
    "cash ebitda margin (term)": "ebitda_margin_terminal",
    "adj ebitda margin": "ebitda_margin_terminal",
    "run-rate ebitda margin": "ebitda_margin_terminal",
    "ebitdax margin": "ebitda_margin_terminal",
    "brand ebitda margin": "ebitda_margin_terminal",

    "d&a % of revenue": "da_pct_rev",
    "d&a % rev": "da_pct_rev",
    "d&a % billings": "da_pct_rev",
    "d&a % revenue": "da_pct_rev",

    "capex % of revenue": "capex_pct_rev",
    "capex % rev": "capex_pct_rev",
    "capex % billings": "capex_pct_rev",

    "nwc change % of revenue": "nwc_pct_rev",
    "nwc % rev": "nwc_pct_rev",

    "tax rate": "tax_rate",
    "wacc": "wacc",
    "terminal growth rate": "terminal_growth_rate",
    "terminal growth": "terminal_growth_rate",
}

COMPANY_NAME_MAP = {
    "medtech solutions inc.": "MedTech Solutions Inc.",
    "buildright construction holdings": "BuildRight Construction",
    "cloudsync saas platform": "CloudSync SaaS Platform",
    "premier retail holdings": "Premier Retail Holdings",
    "apex industrial manufacturing corp": "Apex Industrial Manufacturing",
    "apex industrial manufacturing": "Apex Industrial Manufacturing",
    "natfresh food & beverage distribution": "NatFresh Food & Beverage",
    "natfresh food & beverage": "NatFresh Food & Beverage",
    "fasttrack logistics & transport llc": "FastTrack Logistics",
    "fasttrack logistics": "FastTrack Logistics",
    "carefirst healthcare services group": "CareFirst Healthcare Services",
    "carefirst healthcare services": "CareFirst Healthcare Services",
    "permian energy services partners lp": "Permian Energy Services",
    "permian energy services": "Permian Energy Services",
    "vitality consumer brands inc.": "Vitality Consumer Brands",
    "vitality consumer brands": "Vitality Consumer Brands",
}

PL_SHEET_NAMES = {
    "p&l summary", "p&l", "income statement", "segment p&l",
    "monthly p&l", "financial summary", "financial results", "p&l by channel"
}

KPI_SHEET_NAMES = {
    "operating kpis", "operational kpis", "saas metrics", "category mix",
    "fleet & ops kpis", "operational data", "backlog & pipeline",
    "store kpis", "clinical kpis", "brand kpis"
}

DCF_SHEET_NAMES = {"dcf model", "dcf", "valuation"}

PERIOD_TOKENS = ("fy", "q1", "q2", "q3", "q4", "jan", "feb", "mar", "apr",
                 "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec")

REVENUE_LABELS = [
    "total revenue", "total net revenue", "net revenue", "contract billings",
    "net sales", "net patient revenue",
]

EBITDA_LABELS = [
    "adjusted ebitda", "adj. ebitda", "cash ebitda", "brand ebitda",
    "ebitdax", "run-rate ebitda", "ebitda",
]


# ================================================================================
# UTILITY FUNCTIONS
# ================================================================================

def classify_sheet(sheet_name: str) -> str:
    """Classify sheet as 'pl', 'kpi', 'dcf', or 'unknown'."""
    lower = sheet_name.lower()
    if lower in PL_SHEET_NAMES:
        return "pl"
    if lower in KPI_SHEET_NAMES:
        return "kpi"
    if lower in DCF_SHEET_NAMES:
        return "dcf"
    return "unknown"


def is_numeric(val) -> bool:
    """Return True if val is a real number (not formula, not None)."""
    if val is None:
        return False
    if isinstance(val, (int, float)):
        return True
    return False


def get_unit_scale(filename: str, subtitle_row: Optional[str]) -> float:
    """Return multiplier to convert raw values to $MM."""
    subtitle_lower = (subtitle_row or "").lower()
    # Permian is $MM
    if "$mm" in subtitle_lower or "all figures $mm" in subtitle_lower:
        return 1.0
    # Everyone else is $000s → multiply by 0.001 to get $MM
    return 0.001


def normalize_period(raw: str) -> Optional[str]:
    """Normalize period string to canonical format, or None if skip."""
    s = raw.strip()
    slow = s.lower()

    # Skip non-period columns
    skip_words = ("yoy", "qoq", "variance", "budget", "ytd", "chg", "growth", "pct", "%")
    if any(w in slow for w in skip_words):
        return None

    # Skip formula strings
    if s.startswith("="):
        return None

    # FY20xxA or FY20xx
    m = re.match(r"FY(20\d\d)([AE]?)$", s, re.IGNORECASE)
    if m:
        yr, suffix = m.group(1), (m.group(2).upper() or "A")
        return f"FY{yr}{suffix}"

    # "Actual Q1'25" → Q1-2025A
    m = re.match(r"Actual\s+Q([1-4])['\s](\d{2})$", s, re.IGNORECASE)
    if m:
        q, yy = m.group(1), int(m.group(2))
        yr = 2000 + yy
        return f"Q{q}-{yr}A"

    # Q1'24A or Q1'24 or Q1 2024A
    m = re.match(r"Q([1-4])['\s](20\d{2}|\d{2})[AE]?", s, re.IGNORECASE)
    if m:
        q, yr_raw = m.group(1), m.group(2)
        yr = int(yr_raw) if len(yr_raw) == 4 else 2000 + int(yr_raw)
        return f"Q{q}-{yr}A"

    # "Q1 2024A" with space (explicit FY prefix variant)
    m = re.match(r"Q([1-4])\s+(20\d{2})([AE]?)", s, re.IGNORECASE)
    if m:
        q, yr, suf = m.group(1), m.group(2), (m.group(3).upper() or "A")
        return f"Q{q}-{yr}{suf}"

    # Monthly: "Jan'25"
    MONTHS = {
        "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
        "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12
    }
    m = re.match(r"([A-Za-z]{3})['\s](\d{2})$", s)
    if m:
        mon, yy = m.group(1).lower(), int(m.group(2))
        if mon in MONTHS:
            yr = 2000 + yy
            return f"{mon.capitalize()}-{yr}"

    return None


def infer_period_type(period: str) -> str:
    """Infer period type from normalized period string."""
    if period.startswith("FY"):
        return "estimate" if period.endswith("E") else "annual"
    if re.match(r"Q\d-\d{4}", period):
        return "quarterly"
    if re.match(r"[A-Z][a-z]{2}-\d{4}", period):
        return "monthly"
    return "other"


def find_header_row(ws) -> Tuple[Optional[int], Optional[list]]:
    """Find header row by scanning rows 1-9 for period tokens. Return (1-based row idx, row values) or (None, None)."""
    for row_idx in range(1, 10):
        row = [ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)]
        non_null = [v for v in row if v is not None]
        if len(non_null) < 2:
            continue
        # Check if at least one cell looks like a period
        has_period = any(
            str(v).lower()[:2] in PERIOD_TOKENS or str(v).lower()[:4] == "fy20"
            for v in non_null
        )
        if has_period:
            return row_idx, row
    return None, None


def parse_period_columns(header_row: list) -> Dict[int, str]:
    """Return {col_index_0based: normalized_period} for valid non-skip columns."""
    result = {}
    for idx, val in enumerate(header_row):
        if val is None:
            continue
        s = str(val).strip()
        # Skip formula cells
        if s.startswith("="):
            continue
        norm = normalize_period(s)
        if norm:
            result[idx] = norm
    return result


def find_metric_row(ws, labels: List[str]) -> Tuple[Optional[int], Optional[str]]:
    """Find first metric row matching priority-ordered labels. Return (1-based row idx, matched raw label) or (None, None)."""
    for row_idx in range(1, ws.max_row + 1):
        cell_val = ws.cell(row_idx, 1).value
        if cell_val is None:
            continue
        raw = str(cell_val).strip()
        # Skip indented sub-rows (leading spaces)
        if raw.startswith(" "):
            continue
        # Remove unit suffixes like "($000s)", "($MM)"
        cleaned = re.sub(r"\s*\(.*?\)", "", raw).strip().lower()
        for label in labels:
            if cleaned == label.lower():
                return row_idx, raw
    return None, None


def infer_kpi_unit(label: str) -> str:
    """Infer KPI unit from label."""
    low = label.lower()
    if "%" in label or "margin" in low or "rate" in low or "pct" in low or "utilization" in low or "retention" in low:
        return "pct"
    if "$000" in label:
        return "000s_usd"
    if "$mm" in low:
        return "mm_usd"
    if "($)" in label or "per unit" in low or "cac" in low or "ltv" in low or "cost per" in low:
        return "usd"
    if "count" in low or "fte" in low or "headcount" in low or "rigs" in low or "trucks" in low or "accounts" in low:
        return "count"
    if "(x)" in label or "ratio" in low or "ltv/cac" in low:
        return "x"
    return "raw"


# ================================================================================
# SHEET INGESTION FUNCTIONS
# ================================================================================

def ingest_pl_sheet(ws, company_id: int, filename: str, scale: float) -> List[Dict]:
    """Parse P&L sheet. Return list of financial_metrics records to upsert."""
    records = []

    # Find header row and parse period columns
    header_row_idx, header_row = find_header_row(ws)
    if not header_row_idx:
        return records

    period_cols = parse_period_columns(header_row)  # {col_idx_0based: normalized_period}
    if not period_cols:
        return records

    # Find revenue row
    rev_row, rev_label = find_metric_row(ws, REVENUE_LABELS)

    # Find EBITDA row (priority-ordered)
    ebitda_row, ebitda_label = find_metric_row(ws, EBITDA_LABELS)

    # Extract values per period
    for col_0, period in period_cols.items():
        col_1 = col_0 + 1  # openpyxl is 1-based

        if rev_row:
            val = ws.cell(rev_row, col_1).value
            if is_numeric(val):
                records.append({
                    "company_id": company_id,
                    "period": period,
                    "metric_name": "revenue",
                    "raw_label": rev_label,
                    "value": round(float(val) * scale, 4),
                    "unit": "mm_usd",
                    "period_type": infer_period_type(period),
                    "source_file": filename,
                })

        if ebitda_row:
            val = ws.cell(ebitda_row, col_1).value
            if is_numeric(val):
                # Map to canonical name
                canonical = CANONICAL_METRIC_MAP.get(
                    re.sub(r"\s*\(.*?\)", "", ebitda_label).strip().lower(), "ebitda"
                )
                records.append({
                    "company_id": company_id,
                    "period": period,
                    "metric_name": canonical,
                    "raw_label": ebitda_label,
                    "value": round(float(val) * scale, 4),
                    "unit": "mm_usd",
                    "period_type": infer_period_type(period),
                    "source_file": filename,
                })

    return records


def ingest_kpi_sheet(ws, company_id: int, filename: str) -> List[Dict]:
    """Parse KPI sheet. Return list of company_kpis records."""
    records = []

    # Find header row
    header_row_idx, header_row_vals = find_header_row(ws)
    if not header_row_idx:
        return records

    period_cols = parse_period_columns(header_row_vals)
    if not period_cols:
        return records

    # Parse data rows
    for ridx in range(header_row_idx + 1, ws.max_row + 1):
        kpi_name = ws.cell(ridx, 1).value
        if not kpi_name or str(kpi_name).startswith("="):
            continue
        kpi_name = str(kpi_name).strip()
        # Skip metadata rows
        if not kpi_name or kpi_name.upper() in ("KPI", "METRIC", "UNIT ECONOMICS"):
            continue

        unit = infer_kpi_unit(kpi_name)

        for col_0, period in period_cols.items():
            val = ws.cell(ridx, col_0 + 1).value
            if not is_numeric(val):
                continue
            records.append({
                "company_id": company_id,
                "period": period,
                "kpi_name": kpi_name,
                "kpi_value": float(val),
                "kpi_unit": unit,
                "source_file": filename,
            })

    return records


def ingest_dcf_sheet(ws, company_id: int, filename: str) -> List[Dict]:
    """Parse DCF sheet. Return list of dcf_assumptions records."""
    records = []

    for ridx in range(1, min(16, ws.max_row + 1)):
        label_val = ws.cell(ridx, 1).value
        num_val = ws.cell(ridx, 2).value
        if not label_val or not is_numeric(num_val):
            continue
        label = str(label_val).strip()
        canonical = DCF_ASSUMPTION_MAP.get(label.lower())
        if canonical:
            records.append({
                "company_id": company_id,
                "assumption_name": canonical,
                "raw_label": label,
                "value": float(num_val),
                "source_file": filename,
            })

    return records


# ================================================================================
# BULK INGESTION
# ================================================================================

def ingest_workbook(file_input, conn, filename: str = None) -> Dict:
    """Ingest all sheets of one workbook. file_input can be a file path (str) or file-like object (BytesIO).
    Return result dict with detailed status."""
    try:
        wb = openpyxl.load_workbook(file_input, data_only=True)
    except Exception as e:
        return {"success": False, "error": f"Failed to load workbook: {str(e)}"}

    # Determine filename for source_file tracking
    if filename is None:
        if isinstance(file_input, str):
            filename = os.path.basename(file_input)
        else:
            # File-like object - try to get name attribute
            filename = getattr(file_input, 'name', 'uploaded_file.xlsx')
            if '/' in filename or '\\' in filename:
                filename = os.path.basename(filename)

    # Get company name from row 1, col A - with validation
    first_ws = wb[wb.sheetnames[0]]
    raw_name = str(first_ws.cell(1, 1).value or "").strip()

    if not raw_name:
        return {"success": False, "error": "No company name found in Row 1, Column A. Please ensure the file has the company name in the first cell."}

    # Reject if it looks like a metric/data row (contains $ or % or common metric keywords)
    metric_keywords = ["revenue", "ebitda", "margin", "cost", "expense", "growth", "multiple", "valuation"]
    if any(kw in raw_name.lower() for kw in metric_keywords):
        return {"success": False, "error": f"Row 1 contains '{raw_name}' which looks like a metric, not a company name. Ensure company name is in Row 1, Column A."}

    # Try to find canonical name in map, otherwise use raw name as-is for new companies
    canonical_name = COMPANY_NAME_MAP.get(raw_name.lower(), raw_name)

    if not canonical_name:
        return {"success": False, "error": f"Invalid company name: {raw_name}"}

    # Look up or create company in DB
    try:
        company_id = conn.execute(
            sqla_text("SELECT id FROM companies WHERE name = :name"),
            {"name": canonical_name}
        ).scalar()

        if not company_id:
            # Auto-create new company with defaults
            result = conn.execute(
                sqla_text("INSERT INTO companies(name, sector, entry_year, entry_ev_mm, ownership_pct) "
                         "VALUES(:name, :sector, :entry_year, :entry_ev_mm, :ownership_pct) "
                         "RETURNING id"),
                {"name": canonical_name, "sector": "Unknown", "entry_year": 2024,
                 "entry_ev_mm": 0.0, "ownership_pct": 0.0}
            )
            company_id = result.scalar()
    except Exception as e:
        return {"success": False, "error": f"DB error: {str(e)}"}

    # Clear old data for this company+source_file
    try:
        conn.execute(
            sqla_text("DELETE FROM financial_metrics WHERE company_id = :id AND source_file = :f"),
            {"id": company_id, "f": filename}
        )
        conn.execute(
            sqla_text("DELETE FROM company_kpis WHERE company_id = :id AND source_file = :f"),
            {"id": company_id, "f": filename}
        )
        conn.execute(
            sqla_text("DELETE FROM dcf_assumptions WHERE company_id = :id AND source_file = :f"),
            {"id": company_id, "f": filename}
        )
    except Exception as e:
        return {"success": False, "error": f"Failed to clear old data: {str(e)}"}

    # Parse all sheets
    subtitle_row = str(first_ws.cell(2, 1).value or "")
    scale = get_unit_scale(filename, subtitle_row)

    fin_records = []
    kpi_records = []
    dcf_records = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_type = classify_sheet(sheet_name)
        if sheet_type == "pl":
            fin_records += ingest_pl_sheet(ws, company_id, filename, scale)
        elif sheet_type == "kpi":
            kpi_records += ingest_kpi_sheet(ws, company_id, filename)
        elif sheet_type == "dcf":
            dcf_records += ingest_dcf_sheet(ws, company_id, filename)

    # Bulk insert/upsert
    try:
        for rec in fin_records:
            conn.execute(
                sqla_text("""
                    INSERT INTO financial_metrics
                    (company_id, period, metric_name, raw_label, value, unit, period_type, source_file)
                    VALUES (:company_id, :period, :metric_name, :raw_label, :value, :unit, :period_type, :source_file)
                    ON CONFLICT(company_id, period, metric_name) DO UPDATE
                    SET value=excluded.value, raw_label=excluded.raw_label
                """),
                rec
            )

        for rec in kpi_records:
            conn.execute(
                sqla_text("""
                    INSERT INTO company_kpis
                    (company_id, period, kpi_name, kpi_value, kpi_unit, source_file)
                    VALUES (:company_id, :period, :kpi_name, :kpi_value, :kpi_unit, :source_file)
                    ON CONFLICT(company_id, period, kpi_name) DO UPDATE
                    SET kpi_value=excluded.kpi_value, kpi_unit=excluded.kpi_unit
                """),
                rec
            )

        for rec in dcf_records:
            conn.execute(
                sqla_text("""
                    INSERT INTO dcf_assumptions
                    (company_id, assumption_name, raw_label, value, source_file)
                    VALUES (:company_id, :assumption_name, :raw_label, :value, :source_file)
                    ON CONFLICT(company_id, assumption_name) DO UPDATE
                    SET value=excluded.value, raw_label=excluded.raw_label
                """),
                rec
            )
    except Exception as e:
        return {"success": False, "error": f"Insert failed: {str(e)}"}

    return {
        "success": True,
        "company": canonical_name,
        "company_id": company_id,
        "financials_ingested": len(fin_records),
        "kpis_ingested": len(kpi_records),
        "dcf_assumptions_ingested": len(dcf_records),
    }


def bulk_ingest_all(portco_dir: str = None):
    """Bulk ingest all files from Portco data folder. Skip if data already exists."""
    if portco_dir is None:
        portco_dir = r"C:\AI\Database_Agent\Portco data"

    from database import engine

    # Check if already ingested
    with engine.connect() as conn:
        count = conn.execute(
            sqla_text("SELECT COUNT(*) FROM financial_metrics")
        ).scalar()
        if count and count > 0:
            print(f"  Financial data already ingested ({count} records). Skipping bulk ingest.")
            return []

    # Ingest all files
    if not os.path.isdir(portco_dir):
        print(f"  Portco data directory not found: {portco_dir}")
        return []

    files = sorted(f for f in os.listdir(portco_dir) if f.endswith(".xlsx"))
    results = []

    with engine.begin() as conn:
        for fname in files:
            path = os.path.join(portco_dir, fname)
            result = ingest_workbook(path, conn)
            results.append(result)
            if result["success"]:
                print(f"  {fname}: OK ({result['financials_ingested']} fin, {result['kpis_ingested']} kpi, {result['dcf_assumptions_ingested']} dcf)")
            else:
                print(f"  {fname}: FAIL - {result.get('error')}")

    return results


if __name__ == "__main__":
    # For standalone testing
    results = bulk_ingest_all()
    print(f"\nIngested {len(results)} files")
