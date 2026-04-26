"""
app.py — Portfolio Analysis Agent Web UI
Run: py app.py
Open: http://localhost:5000
"""

import json
import logging
import os
import queue
import threading
import time

import litellm
from datetime import datetime
from flask import Flask, render_template_string, request, Response, jsonify
from dotenv import load_dotenv
from database import (execute_query, get_schema, seed_sample_data, engine,
                      upsert_company, upsert_financials, upsert_kpi,
                      get_company_detail, get_company_comps,
                      upsert_thesis, get_thesis, upsert_milestone,
                      delete_milestone, get_all_thesis_statuses,
                      get_value_creation, get_portfolio_value_creation)
from tools import TOOL_DEFINITIONS, execute_tool
from prompts import SYSTEM_PROMPT
from sqlalchemy import text

load_dotenv()
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
MODEL = os.getenv("MODEL", "anthropic/claude-sonnet-4-6")
litellm.set_verbose = False

def _log_env_status():
    env_keys = ["ANTHROPIC_API_KEY", "OPENAI_API_KEY", "GOOGLE_API_KEY"]
    status = {key: bool(os.getenv(key)) for key in env_keys}
    logger.info("Litellm startup check: MODEL=%s, env_keys=%s", MODEL, status)


_log_env_status()


def _serialize_tool_call(tool_call):
    """Convert ChatCompletionMessageToolCall to a plain dict."""
    try:
        function = getattr(tool_call, "function", None)
        func_dict = {}
        if function:
            func_dict["name"] = str(getattr(function, "name", ""))
            func_dict["arguments"] = str(getattr(function, "arguments", ""))
        
        return {
            "id": str(getattr(tool_call, "id", "")),
            "type": str(getattr(tool_call, "type", "function")),
            "function": func_dict,
        }
    except Exception as e:
        logger.warning("Failed to serialize tool_call: %s", e)
        return {"id": "", "type": "function", "function": {"name": "", "arguments": ""}}


def _parse_tool_args(tool_call):
    function = getattr(tool_call, "function", None)
    if function is None:
        return {}
    raw_args = getattr(function, "arguments", "")
    if isinstance(raw_args, dict):
        return raw_args
    if not isinstance(raw_args, str):
        return {}
    try:
        return json.loads(raw_args)
    except json.JSONDecodeError:
        return {}


def _get_provider_api_key():
    provider = MODEL.split("/")[0].lower() if MODEL else ""
    env_map = {
        "anthropic": "ANTHROPIC_API_KEY",
        "openai": "OPENAI_API_KEY",
        "google": "GOOGLE_API_KEY",
        "gpt": "OPENAI_API_KEY",
    }
    key_name = env_map.get(provider)
    if key_name:
        logger.info("Provider %s mapped to env var %s", provider, key_name)
        return os.getenv(key_name)
    logger.info("Provider %s has no mapped env var", provider)
    return None


# ── Agent runner ───────────────────────────────────────────────────────────────
def run_agent_streaming(user_question: str, response_queue: queue.Queue):
    messages = [
        {"role": "system",  "content": SYSTEM_PROMPT},
        {"role": "user",    "content": user_question},
    ]
    response_queue.put({"type": "status", "msg": f"Thinking..."})

    for iteration in range(12):
        try:
            response = litellm.completion(
                model=MODEL,
                messages=messages,
                tools=TOOL_DEFINITIONS,
                tool_choice="auto",
                max_tokens=4096,
                api_key=_get_provider_api_key(),
            )
        except Exception as e:
            response_queue.put({"type": "error", "msg": str(e)})
            return

        message = response.choices[0].message
        
        # Safely convert tool_calls iterator to list
        try:
            tool_calls_list = list(message.tool_calls) if message.tool_calls else []
        except (TypeError, AttributeError):
            tool_calls_list = []
        
        messages.append({
            "role": "assistant",
            "content": message.content or "",
            "tool_calls": [_serialize_tool_call(tc) for tc in tool_calls_list],
        })

        if not tool_calls_list:
            response_queue.put({"type": "answer", "msg": message.content or ""})
            return

        for tc in tool_calls_list:
            name = tc.function.name
            args = _parse_tool_args(tc)
            logger.info(
                "Received tool call: name=%s id=%s args=%s",
                name,
                getattr(tc, "id", None),
                getattr(getattr(tc, "function", None), "arguments", None),
            )

            response_queue.put({"type": "tool_start", "tool": name,
                                 "detail": args.get("rationale") or args.get("sql","")[:80]})

            result = execute_tool(name, args)

            # If it's a query result, send the data for table rendering
            if name == "execute_sql_query":
                try:
                    parsed = json.loads(result)
                    if parsed.get("success") and parsed.get("rows"):
                        response_queue.put({"type": "table",
                                            "columns": parsed["columns"],
                                            "rows":    parsed["rows"][:50],
                                            "count":   parsed["row_count"]})
                except:
                    pass

            response_queue.put({"type": "tool_done", "tool": name,
                                 "rows": json.loads(result).get("row_count","") if name=="execute_sql_query" else ""})

            messages.append({"role":"tool","tool_call_id":tc.id,"name":name,"content":result})

    response_queue.put({"type":"error","msg":"Max iterations reached."})


# ── Data helpers ───────────────────────────────────────────────────────────────
def get_portfolio_summary():
    q = execute_query("""
        SELECT c.id, c.name, c.sector, qf.revenue_mm, qf.ebitda_mm, qf.ebitda_margin, qf.yoy_growth
        FROM companies c
        JOIN quarterly_financials qf ON c.id = qf.company_id
        WHERE qf.period = '2024-Q4'
        ORDER BY qf.ebitda_margin DESC
    """)
    return q

def get_comps_summary():
    q = execute_query("""
        SELECT ticker, name, sector_key, market_cap_mm, ev_ebitda, ebitda_margin, revenue_growth, last_updated
        FROM market_comps
        ORDER BY sector_key, market_cap_mm DESC NULLS LAST
        LIMIT 50
    """)
    return q

def get_comp_count():
    try:
        with engine.connect() as conn:
            return conn.execute(text("SELECT COUNT(*) FROM market_comps")).scalar()
    except:
        return 0

def get_price_history(ticker):
    q = execute_query(f"""
        SELECT date, close FROM comp_price_history
        WHERE ticker = '{ticker}'
        ORDER BY date ASC
    """)
    return q


# ── HTML ───────────────────────────────────────────────────────────────────────
HTML = r'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>PE Portfolio Intelligence</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Sora:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
<script src="https://cdnjs.cloudflare.com/ajax/libs/marked/9.1.6/marked.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
:root {
  --navy:  #0A1628;
  --navy2: #0F1F3D;
  --blue:  #1E6FD9;
  --teal:  #0ABFBC;
  --gold:  #F5A623;
  --green: #27AE60;
  --red:   #E74C3C;
  --grey1: #F8F9FC;
  --grey2: #EEF0F5;
  --grey3: #B0B8CC;
  --text:  #1A1F36;
  --white: #FFFFFF;
  --sidebar-w: 280px;
  --header-h:  60px;
}
*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: "Sora", sans-serif; background: var(--grey1); color: var(--text); height: 100vh; display: flex; flex-direction: column; overflow: hidden; }

/* Header */
.header { height: var(--header-h); background: var(--navy); display: flex; align-items: center; padding: 0 24px; gap: 16px; border-bottom: 2px solid var(--gold); flex-shrink: 0; z-index: 50; }
.header-logo { width: 36px; height: 36px; background: var(--gold); border-radius: 8px; display: flex; align-items: center; justify-content: center; font-size: 18px; flex-shrink: 0; }
.header-title { font-size: 18px; font-weight: 700; color: white; letter-spacing: -0.3px; }
.header-sub { font-size: 11px; color: var(--grey3); letter-spacing: 1px; text-transform: uppercase; }
.header-spacer { flex: 1; }
.header-model { font-family: "JetBrains Mono", monospace; font-size: 11px; color: var(--teal); background: rgba(10,191,188,0.1); border: 1px solid rgba(10,191,188,0.3); padding: 4px 10px; border-radius: 20px; }
.header-tabs { display: flex; gap: 8px; }
.tab-btn { padding: 8px 16px; border: none; border-radius: 8px; font-family: "Sora", sans-serif; font-size: 13px; font-weight: 500; cursor: pointer; transition: all 0.2s; background: rgba(255,255,255,0.07); color: var(--grey3); }
.tab-btn.active { background: var(--blue); color: white; }

/* Layout */
.main { display: flex; flex: 1; overflow: hidden; }

/* Sidebar */
.sidebar { width: var(--sidebar-w); background: var(--navy2); border-right: 1px solid rgba(255,255,255,0.06); display: flex; flex-direction: column; flex-shrink: 0; overflow: hidden; }
.sidebar-section { padding: 16px 20px 10px; font-size: 11px; font-weight: 600; letter-spacing: 1.2px; text-transform: uppercase; color: var(--grey3); }
.sidebar-item { padding: 10px 20px; font-size: 13px; color: rgba(255,255,255,0.7); cursor: pointer; display: flex; align-items: center; gap: 8px; transition: all 0.15s; border-left: 3px solid transparent; }
.sidebar-item:hover { background: rgba(255,255,255,0.05); color: white; }
.sidebar-item.active { background: rgba(30,111,217,0.15); color: var(--teal); border-left-color: var(--teal); }
.sidebar-item .sector-dot { width: 8px; height: 8px; border-radius: 50%; flex-shrink: 0; }
.sidebar-item .margin-pill { margin-left: auto; font-family: "JetBrains Mono", monospace; font-size: 11px; color: var(--teal); }
.sidebar-divider { height: 1px; background: rgba(255,255,255,0.06); margin: 8px 0; }
.comp-count { padding: 10px 20px; font-size: 12px; color: var(--grey3); }
.refresh-btn { margin: 10px 20px; padding: 10px 14px; background: rgba(245,166,35,0.1); border: 1px solid rgba(245,166,35,0.3); color: var(--gold); border-radius: 8px; font-family: "Sora",sans-serif; font-size: 12px; font-weight: 500; cursor: pointer; text-align: center; transition: all 0.2s; }
.refresh-btn:hover { background: rgba(245,166,35,0.2); }

/* Content panels */
.content { flex: 1; display: flex; flex-direction: column; overflow: hidden; }
.panel { flex: 1; overflow: hidden; display: none; flex-direction: column; }
.panel.active { display: flex; }

/* Dashboard panel */
.dashboard { flex: 1; overflow-y: auto; padding: 24px; }
.dashboard-section { margin-bottom: 32px; }
.dashboard-title { font-size: 20px; font-weight: 700; color: var(--navy); margin-bottom: 16px; }
.dashboard-subtitle { font-size: 14px; color: var(--grey3); margin-bottom: 16px; }
.data-table-wrap { overflow-x: auto; margin-bottom: 16px; border-radius: 8px; border: 1px solid var(--grey2); }
.data-table { border-collapse: collapse; font-size: 13px; width: 100%; }
.data-table th { background: var(--navy); color: white; padding: 10px 14px; text-align: left; font-size: 12px; font-weight: 600; letter-spacing: 0.3px; white-space: nowrap; }
.data-table td { padding: 8px 14px; border-bottom: 1px solid var(--grey2); white-space: nowrap; }
.data-table tr:hover td { background: rgba(30,111,217,0.04); }
.data-table-count { font-size: 12px; color: var(--grey3); margin-top: 8px; padding: 0 4px; }
.metric-card { background: white; border-radius: 12px; padding: 20px; margin-bottom: 16px; box-shadow: 0 2px 12px rgba(0,0,0,0.08); }
.metric-title { font-size: 14px; font-weight: 600; color: var(--navy); margin-bottom: 8px; }
.metric-value { font-size: 24px; font-weight: 700; color: var(--blue); }
.metric-sub { font-size: 12px; color: var(--grey3); margin-top: 4px; }

/* Summary stats row */
.summary-stats { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 16px; margin-bottom: 32px; }
.stat-card { background: white; border-radius: 12px; padding: 20px; box-shadow: 0 2px 12px rgba(0,0,0,0.08); border-top: 4px solid var(--blue); }
.stat-card.top-performer { border-top-color: var(--green); }
.stat-label { font-size: 12px; font-weight: 600; letter-spacing: 0.5px; text-transform: uppercase; color: var(--grey3); margin-bottom: 8px; }
.stat-value { font-size: 28px; font-weight: 700; color: var(--navy); }
.stat-sub { font-size: 12px; color: var(--grey3); margin-top: 6px; }

/* Company cards grid */
.cards-header { display: flex; justify-content: space-between; align-items: center; margin-bottom: 16px; gap: 12px; }
.sort-dropdown { padding: 8px 12px; border: 1px solid var(--grey2); border-radius: 8px; font-size: 13px; background: white; cursor: pointer; font-family: "Sora", sans-serif; }
.company-cards { display: grid; grid-template-columns: repeat(auto-fill, minmax(280px, 1fr)); gap: 16px; }
.company-card { background: white; border-radius: 12px; padding: 20px; box-shadow: 0 2px 12px rgba(0,0,0,0.08); transition: all 0.2s; cursor: pointer; border: 1px solid var(--grey2); }
.company-card:hover { transform: translateY(-2px); box-shadow: 0 4px 16px rgba(0,0,0,0.12); }
.company-card-header { display: flex; justify-content: space-between; align-items: start; margin-bottom: 12px; }
.company-name { font-size: 16px; font-weight: 700; color: var(--navy); margin-bottom: 4px; }
.sector-badge { display: inline-block; padding: 4px 10px; background: rgba(30,111,217,0.1); color: var(--blue); border-radius: 6px; font-size: 11px; font-weight: 600; }
.company-metrics { display: grid; grid-template-columns: 1fr 1fr; gap: 12px; margin-bottom: 12px; }
.metric-item { padding: 10px; background: var(--grey1); border-radius: 8px; }
.metric-item-label { font-size: 11px; color: var(--grey3); margin-bottom: 4px; font-weight: 500; }
.metric-item-value { font-size: 18px; font-weight: 700; color: var(--navy); }
.metric-item-value.positive { color: var(--green); }
.metric-item-value.negative { color: var(--red); }
.company-card-chart { height: 40px; margin-top: 12px; border-top: 1px solid var(--grey2); padding-top: 12px; }

/* Company detail modal */
.modal { display: none; position: fixed; z-index: 100; left: 0; top: 0; width: 100%; height: 100%; background: rgba(0,0,0,0.5); }
.modal.show { display: flex; }
.modal-content { background: white; margin: auto; border-radius: 12px; max-width: 900px; max-height: 90vh; overflow-y: auto; width: 95%; box-shadow: 0 10px 40px rgba(0,0,0,0.2); }
.modal-header { padding: 24px 28px; border-bottom: 1px solid var(--grey2); display: flex; justify-content: space-between; align-items: center; }
.modal-title { font-size: 24px; font-weight: 700; color: var(--navy); }
.modal-close { font-size: 24px; cursor: pointer; color: var(--grey3); border: none; background: none; }
.modal-close:hover { color: var(--navy); }
.modal-body { padding: 28px; }
.modal-section { margin-bottom: 28px; }
.modal-section-title { font-size: 16px; font-weight: 700; color: var(--navy); margin-bottom: 16px; }
.modal-section-table { width: 100%; border-collapse: collapse; font-size: 13px; }
.modal-section-table th { background: var(--grey1); padding: 10px 12px; text-align: left; font-weight: 600; color: var(--navy); border-bottom: 1px solid var(--grey2); }
.modal-section-table td { padding: 10px 12px; border-bottom: 1px solid var(--grey2); }
.modal-section-table tr:hover td { background: rgba(30,111,217,0.02); }
.chart-container { position: relative; height: 300px; margin-bottom: 20px; }

/* Chat panel */
.chat-messages { flex: 1; overflow-y: auto; padding: 20px; display: flex; flex-direction: column; gap: 16px; }
.chat-empty { text-align: center; margin: auto; color: var(--grey3); }
.chat-empty-icon { font-size: 48px; margin-bottom: 16px; }
.chat-empty-title { font-size: 18px; font-weight: 600; color: var(--text); margin-bottom: 8px; }
.chat-empty-sub { font-size: 13px; line-height: 1.6; max-width: 380px; }
.suggestions { display: flex; flex-wrap: wrap; gap: 8px; justify-content: center; margin-top: 20px; }
.suggestion { padding: 8px 14px; background: white; border: 1.5px solid var(--grey2); border-radius: 20px; font-size: 12px; color: var(--text); cursor: pointer; transition: all 0.2s; }
.suggestion:hover { border-color: var(--blue); color: var(--blue); background: rgba(30,111,217,0.04); }

.msg { display: flex; flex-direction: column; gap: 4px; max-width: 820px; }
.msg.user { align-self: flex-end; align-items: flex-end; }
.msg.agent { align-self: flex-start; }
.msg-label { font-size: 10px; font-weight: 600; letter-spacing: 0.8px; text-transform: uppercase; color: var(--grey3); padding: 0 4px; }
.msg-bubble { padding: 14px 18px; border-radius: 16px; font-size: 13.5px; line-height: 1.7; }
.msg.user .msg-bubble { background: var(--blue); color: white; border-bottom-right-radius: 4px; }
.msg.agent .msg-bubble { background: white; color: var(--text); border-bottom-left-radius: 4px; box-shadow: 0 1px 8px rgba(0,0,0,0.06); }
.msg.agent .msg-bubble h2 { font-size: 15px; font-weight: 700; margin: 12px 0 6px; color: var(--navy); }
.msg.agent .msg-bubble h3 { font-size: 13px; font-weight: 600; margin: 10px 0 4px; color: var(--navy); }
.msg.agent .msg-bubble p { margin: 6px 0; }
.msg.agent .msg-bubble ul, .msg.agent .msg-bubble ol { padding-left: 18px; margin: 6px 0; }
.msg.agent .msg-bubble li { margin: 3px 0; }
.msg.agent .msg-bubble strong { color: var(--navy); }
.msg.agent .msg-bubble table { border-collapse: collapse; width: 100%; font-size: 12px; margin: 10px 0; }
.msg.agent .msg-bubble th { background: var(--navy); color: white; padding: 7px 12px; text-align: left; font-weight: 600; }
.msg.agent .msg-bubble td { padding: 6px 12px; border-bottom: 1px solid var(--grey2); }
.msg.agent .msg-bubble tr:hover td { background: var(--grey1); }
.msg.agent .msg-bubble code { font-family: "JetBrains Mono", monospace; font-size: 11px; background: var(--grey2); padding: 1px 5px; border-radius: 4px; }

.tool-trace { background: var(--grey1); border: 1px solid var(--grey2); border-radius: 10px; padding: 10px 14px; margin: 4px 0; font-size: 11px; }
.tool-trace-row { display: flex; align-items: center; gap: 8px; color: var(--grey3); }
.tool-trace-row .tool-name { color: var(--teal); font-family: "JetBrains Mono", monospace; font-weight: 500; }
.tool-trace-row .tool-detail { color: var(--text); opacity: 0.6; flex: 1; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
.tool-spinner { width: 12px; height: 12px; border: 2px solid var(--grey2); border-top-color: var(--teal); border-radius: 50%; animation: spin 0.7s linear infinite; flex-shrink: 0; }
.tool-done-icon { color: var(--green); font-size: 12px; }

.data-table-wrap { overflow-x: auto; margin: 8px 0; border-radius: 8px; border: 1px solid var(--grey2); }
.data-table { border-collapse: collapse; font-size: 12px; width: 100%; }
.data-table th { background: var(--navy); color: white; padding: 8px 12px; text-align: left; font-size: 11px; font-weight: 600; letter-spacing: 0.3px; white-space: nowrap; }
.data-table td { padding: 6px 12px; border-bottom: 1px solid var(--grey2); white-space: nowrap; }
.data-table tr:hover td { background: rgba(30,111,217,0.04); }
.data-table-count { font-size: 11px; color: var(--grey3); margin-top: 4px; padding: 0 4px; }

/* Thinking dots */
.thinking { display: flex; gap: 5px; align-items: center; padding: 14px 18px; background: white; border-radius: 16px; border-bottom-left-radius: 4px; box-shadow: 0 1px 8px rgba(0,0,0,0.06); }
.dot { width: 7px; height: 7px; border-radius: 50%; background: var(--blue); animation: bounce 1.2s ease-in-out infinite; }
.dot:nth-child(2) { animation-delay: 0.2s; }
.dot:nth-child(3) { animation-delay: 0.4s; }

/* Chat input */
.chat-input-wrap { padding: 16px 20px; background: white; border-top: 1px solid var(--grey2); display: flex; gap: 10px; align-items: flex-end; }
.chat-input { flex: 1; padding: 12px 16px; border: 1.5px solid var(--grey2); border-radius: 12px; font-family: "Sora", sans-serif; font-size: 13px; color: var(--text); resize: none; max-height: 120px; min-height: 48px; transition: border-color 0.2s; line-height: 1.5; }
.chat-input:focus { outline: none; border-color: var(--blue); }
.send-btn { width: 44px; height: 44px; background: var(--blue); border: none; border-radius: 12px; cursor: pointer; color: white; font-size: 18px; display: flex; align-items: center; justify-content: center; flex-shrink: 0; transition: all 0.2s; }
.send-btn:hover { background: #1559B5; transform: translateY(-1px); }
.send-btn:disabled { background: var(--grey3); cursor: not-allowed; transform: none; }

/* Comps panel */
.comps-header { padding: 20px 24px 0; display: flex; align-items: center; justify-content: space-between; }
.comps-title { font-size: 18px; font-weight: 700; color: var(--navy); }
.comps-sub { font-size: 12px; color: var(--grey3); margin-top: 2px; }
.fetch-comps-btn { padding: 10px 20px; background: var(--navy); color: white; border: none; border-radius: 10px; font-family: "Sora",sans-serif; font-size: 13px; font-weight: 600; cursor: pointer; display: flex; align-items: center; gap: 8px; transition: all 0.2s; }
.fetch-comps-btn:hover { background: var(--blue); }
.fetch-comps-btn:disabled { background: var(--grey3); cursor: not-allowed; }
.comps-table-wrap { flex: 1; overflow: auto; padding: 20px 24px; }
.comps-table { width: 100%; border-collapse: collapse; font-size: 12.5px; }
.comps-table th { background: var(--navy); color: white; padding: 10px 14px; text-align: left; font-size: 11px; font-weight: 600; letter-spacing: 0.5px; text-transform: uppercase; position: sticky; top: 0; z-index: 1; }
.comps-table td { padding: 9px 14px; border-bottom: 1px solid var(--grey2); }
.comps-table tr:hover td { background: rgba(30,111,217,0.03); }
.comps-table .sector-badge { display: inline-block; padding: 2px 8px; border-radius: 10px; font-size: 10px; font-weight: 600; background: rgba(10,191,188,0.1); color: var(--teal); }
.comps-table .num { font-family: "JetBrains Mono", monospace; font-size: 12px; text-align: right; }
.comps-table .positive { color: var(--green); }
.comps-table .negative { color: var(--red); }
.comps-empty { text-align: center; padding: 60px; color: var(--grey3); }
.fetch-log { background: var(--navy); color: #90CAF9; font-family: "JetBrains Mono", monospace; font-size: 11px; padding: 12px 16px; border-radius: 10px; max-height: 180px; overflow-y: auto; margin: 0 24px 16px; display: none; line-height: 1.8; }
.fetch-log.show { display: block; }

@keyframes spin { to { transform: rotate(360deg); } }
@keyframes bounce { 0%,80%,100%{transform:scale(0.6)}40%{transform:scale(1)} }
</style>
</head>
<body>

<div class="header">
  <div class="header-logo">📊</div>
  <div>
    <div class="header-title">PE Portfolio Intelligence</div>
    <div class="header-sub">AI-Powered Financial Analysis</div>
  </div>
  <div class="header-spacer"></div>
  <div class="header-tabs">
    <button class="tab-btn active" onclick="switchTab('dashboard')">📈 Dashboard</button>
    <button class="tab-btn" onclick="switchTab('chat')">💬 Chat</button>
    <button class="tab-btn" onclick="switchTab('comps')">📊 Market Comps</button>
    <button class="tab-btn" onclick="switchTab('upload')">📤 Upload</button>
  </div>
  <div class="header-model" id="modelBadge">Loading...</div>
</div>

<div class="main">

  <!-- Sidebar -->
  <div class="sidebar">
    <div class="sidebar-section">Portfolio Companies</div>
    <div id="companyList">
      <div style="padding:20px;color:var(--grey3);font-size:12px">Loading...</div>
    </div>
    <div class="sidebar-divider"></div>
    <div class="sidebar-section">Tools</div>
    <div style="padding:12px 20px;font-size:12px;color:var(--grey3)">Use the Market Comparables tab to analyze your companies against sector peers.</div>
  </div>

  <!-- Content -->
  <div class="content">

    <!-- Dashboard panel -->
    <div class="panel active" id="panel-dashboard">
      <div class="dashboard">
        <div class="dashboard-section">
          <div class="dashboard-title">Portfolio Overview</div>
          <div class="dashboard-subtitle">Company performance and key metrics</div>

          <!-- Summary Stats -->
          <div class="summary-stats" id="summaryStats">
            <div class="stat-card"><div class="stat-label">Total Revenue</div><div class="stat-value">—</div><div class="stat-sub">2024-Q4</div></div>
            <div class="stat-card"><div class="stat-label">Avg EBITDA Margin</div><div class="stat-value">—</div><div class="stat-sub">% of revenue</div></div>
            <div class="stat-card"><div class="stat-label"># Companies</div><div class="stat-value">—</div></div>
            <div class="stat-card top-performer"><div class="stat-label">Top Performer</div><div class="stat-value" style="font-size:18px">—</div><div class="stat-sub">Highest margin</div></div>
          </div>

          <!-- Company Cards with Sorting -->
          <div class="cards-header">
            <div style="font-size:13px;color:var(--grey3)">Select sort order:</div>
            <select class="sort-dropdown" id="sortDropdown" onchange="sortAndRenderCards(this.value)">
              <option value="margin-desc">EBITDA Margin (High to Low)</option>
              <option value="revenue-desc">Revenue (High to Low)</option>
              <option value="growth-desc">YoY Growth (High to Low)</option>
              <option value="name">Alphabetical</option>
            </select>
          </div>

          <!-- Company Cards -->
          <div class="company-cards" id="companyCards">
            <div style="padding:40px;color:var(--grey3);text-align:center;grid-column:1/-1">Loading company data...</div>
          </div>
        </div>

        <div class="dashboard-section">
          <div class="dashboard-title">Market Comparables</div>
          <div class="dashboard-subtitle">Public company data for benchmarking</div>
          <div class="data-table-wrap" id="compsTableWrap">
            <div style="padding:40px;color:var(--grey3);text-align:center">Loading comps data...</div>
          </div>
        </div>
      </div>
    </div>

    <!-- Company Detail Modal -->
    <div class="modal" id="companyModal">
      <div class="modal-content">
        <div class="modal-header">
          <div>
            <div class="modal-title" id="modalCompanyName">—</div>
            <div style="color:var(--grey3);font-size:13px;margin-top:4px"><span class="sector-badge" id="modalSector">—</span></div>
          </div>
          <button class="modal-close" onclick="closeModal()">✕</button>
        </div>
        <div class="modal-body">
          <!-- Current Performance -->
          <div class="modal-section">
            <div class="modal-section-title">Current Performance (2024-Q4)</div>
            <table class="modal-section-table">
              <tr><td style="font-weight:600">Revenue</td><td id="modal-revenue">—</td></tr>
              <tr><td style="font-weight:600">EBITDA</td><td id="modal-ebitda">—</td></tr>
              <tr><td style="font-weight:600">EBITDA Margin</td><td id="modal-margin">—</td></tr>
              <tr><td style="font-weight:600">YoY Growth</td><td id="modal-growth">—</td></tr>
              <tr><td style="font-weight:600">Entry EV</td><td id="modal-entry-ev">—</td></tr>
              <tr><td style="font-weight:600">Ownership %</td><td id="modal-ownership">—</td></tr>
            </table>
          </div>

          <!-- Historical Trend Charts -->
          <div class="modal-section">
            <div class="modal-section-title">Historical Trend (2023-Q1 to 2024-Q4)</div>
            <div class="chart-container"><canvas id="trendChart"></canvas></div>
            <div class="chart-container"><canvas id="marginChart"></canvas></div>
          </div>

          <!-- KPIs -->
          <div class="modal-section" id="kpisSection" style="display:none">
            <div class="modal-section-title">Key Performance Indicators</div>
            <table class="modal-section-table" id="kpisTable">
              <thead><tr><th>Metric</th><th>2024-Q4 Value</th><th>Unit</th></tr></thead>
              <tbody id="kpisTableBody"></tbody>
            </table>
          </div>

          <!-- Peer Benchmarking -->
          <div class="modal-section" id="compsSection" style="display:none">
            <div class="modal-section-title">Peer Benchmarking vs Sector Median</div>
            <table class="modal-section-table" id="compsComparisonTable">
              <thead><tr><th>Metric</th><th>Company</th><th>Sector Median</th><th>vs Median</th></tr></thead>
              <tbody id="compsComparisonBody"></tbody>
            </table>
          </div>
        </div>
      </div>
    </div>

    <!-- Chat panel -->
    <div class="panel" id="panel-chat">
      <div class="chat-messages" id="chatMessages">
        <div class="chat-empty">
          <div class="chat-empty-icon">🔍</div>
          <div class="chat-empty-title">Ask anything about your portfolio</div>
          <div class="chat-empty-sub">The agent queries your live database and public market data to answer your questions.</div>
          <div class="suggestions">
            <div class="suggestion" onclick="sendSuggestion(this)">Which company has the highest EBITDA margin?</div>
            <div class="suggestion" onclick="sendSuggestion(this)">Revenue trend across all companies 2023-2024</div>
            <div class="suggestion" onclick="sendSuggestion(this)">Compare portfolio EBITDA vs public peer medians</div>
            <div class="suggestion" onclick="sendSuggestion(this)">Which sector has the strongest YoY growth?</div>
            <div class="suggestion" onclick="sendSuggestion(this)">Show me the bottom 3 companies by margin</div>
            <div class="suggestion" onclick="sendSuggestion(this)">Permian Energy vs public energy comps</div>
          </div>
        </div>
      </div>
      <div class="chat-input-wrap">
        <textarea class="chat-input" id="chatInput" placeholder="Ask about your portfolio companies, trends, comparables..." rows="1"
          onkeydown="if(event.key==='Enter'&&!event.shiftKey){event.preventDefault();sendMessage()}"
          oninput="this.style.height='auto';this.style.height=this.scrollHeight+'px'"></textarea>
        <button class="send-btn" id="sendBtn" onclick="sendMessage()">↑</button>
      </div>
    </div>

    <!-- Comps panel -->
    <div class="panel" id="panel-comps">
      <div style="display:flex;flex-direction:column;height:100%;overflow:hidden">
        <div class="comps-header" style="border-bottom:1px solid var(--grey2);padding:16px 24px;display:flex;justify-content:space-between;align-items:center">
          <div>
            <div class="comps-title">Market Comparables Analysis</div>
            <div class="comps-sub" id="compsLastUpdated">Select a portfolio company to view sector peers</div>
          </div>
          <button style="background:none;border:none;color:var(--grey3);cursor:pointer;font-size:14px;padding:4px 8px;border-radius:6px;transition:all 0.2s" onclick="toggleFetchPanel()" title="Refresh market data from Yahoo Finance">
            ⚙ Refresh Data
          </button>
        </div>

        <div style="flex:1;overflow-y:auto;padding:24px">
          <!-- Company Selector -->
          <div style="margin-bottom:24px">
            <label style="font-weight:600;color:var(--navy);display:block;margin-bottom:8px">Select Portfolio Company to Compare</label>
            <select id="companySelector" onchange="selectCompanyForComparison(this.value)" style="padding:10px 12px;border:1px solid var(--grey2);border-radius:8px;font-size:13px;width:100%;max-width:400px;background:white;font-family:Sora,sans-serif;cursor:pointer">
              <option value="">-- Choose a company --</option>
            </select>
          </div>

          <!-- Selected Company Card -->
          <div id="selectedCompanyCard" style="display:none;background:white;border-radius:12px;padding:20px;margin-bottom:24px;box-shadow:0 2px 12px rgba(0,0,0,0.08);border-left:4px solid var(--blue)">
            <div style="display:flex;justify-content:space-between;align-items:start;margin-bottom:16px">
              <div>
                <div style="font-size:18px;font-weight:700;color:var(--navy)" id="selectedCompanyName">—</div>
                <div style="color:var(--grey3);font-size:13px;margin-top:4px"><span class="sector-badge" id="selectedCompanySector">—</span></div>
              </div>
              <div style="font-size:11px;color:var(--grey3);text-align:right">Your Portfolio Company<br>(Baseline)</div>
            </div>
            <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(120px,1fr));gap:12px">
              <div style="padding:12px;background:var(--grey1);border-radius:8px">
                <div style="font-size:11px;color:var(--grey3);margin-bottom:4px">Revenue</div>
                <div style="font-size:18px;font-weight:700;color:var(--navy)" id="selectedRevenue">—</div>
              </div>
              <div style="padding:12px;background:var(--grey1);border-radius:8px">
                <div style="font-size:11px;color:var(--grey3);margin-bottom:4px">EBITDA Margin</div>
                <div style="font-size:18px;font-weight:700;color:var(--navy)" id="selectedMargin">—</div>
              </div>
              <div style="padding:12px;background:var(--grey1);border-radius:8px">
                <div style="font-size:11px;color:var(--grey3);margin-bottom:4px">YoY Growth</div>
                <div style="font-size:18px;font-weight:700;color:var(--navy)" id="selectedGrowth">—</div>
              </div>
              <div style="padding:12px;background:var(--grey1);border-radius:8px">
                <div style="font-size:11px;color:var(--grey3);margin-bottom:4px">Entry EV</div>
                <div style="font-size:18px;font-weight:700;color:var(--navy)" id="selectedEV">—</div>
              </div>
            </div>
          </div>

          <!-- Comparables Table -->
          <div id="comparablesSection" style="display:none">
            <div style="font-size:14px;font-weight:600;color:var(--navy);margin-bottom:16px">Sector Comparable Companies</div>
            <div style="overflow-x:auto;border-radius:8px;border:1px solid var(--grey2)">
              <table class="data-table" style="font-size:12px">
                <thead>
                  <tr>
                    <th style="min-width:140px">Company</th>
                    <th style="text-align:right">Ticker</th>
                    <th style="text-align:right">Revenue ($MM)</th>
                    <th style="text-align:right">Growth (%)</th>
                    <th style="text-align:right">EBITDA Margin (%)</th>
                    <th style="text-align:right">EV/EBITDA</th>
                    <th style="text-align:right">EV/Revenue</th>
                  </tr>
                </thead>
                <tbody id="comparablesTableBody">
                </tbody>
              </table>
            </div>
            <div style="font-size:11px;color:var(--grey3);margin-top:12px">
              <strong style="color:var(--navy)">Legend:</strong>
              <span style="color:var(--green);margin-left:8px">Green = Better than your company</span>
              <span style="color:var(--red);margin-left:8px">Red = Worse than your company</span>
              <span style="color:var(--grey3);margin-left:8px">Gray = Not directly comparable</span>
            </div>
          </div>

          <!-- Empty state -->
          <div id="compsEmptyState" style="padding:40px;color:var(--grey3);text-align:center">
            <div style="font-size:48px;margin-bottom:12px">📊</div>
            <div style="font-size:14px;font-weight:600;margin-bottom:8px">No comparables yet</div>
            <div style="font-size:12px">Select a company above to view its sector peers. Make sure to fetch market data first using the button above.</div>
          </div>
        </div>

        <!-- Fetch Log (hidden by default) -->
        <div class="fetch-log" id="fetchLog" style="display:none;padding:16px 24px;border-top:1px solid var(--grey2);max-height:200px;overflow-y:auto"></div>
      </div>
    </div>

    <!-- Upload panel -->
    <div class="panel" id="panel-upload">
      <div style="padding:24px;overflow-y:auto;height:100%">
        <div class="dashboard-title">Upload Portfolio Data</div>
        <div class="dashboard-subtitle">Select an Excel file (.xlsx) to add or update company data</div>

        <div style="background:white;border-radius:12px;padding:24px;margin-bottom:24px;box-shadow:0 2px 12px rgba(0,0,0,0.08)">
          <div style="margin-bottom:16px">
            <label style="font-weight:600;color:var(--navy);display:block;margin-bottom:8px">Select File</label>
            <input type="file" id="uploadFile" accept=".xlsx" style="padding:8px;border:1px solid var(--grey2);border-radius:8px;width:100%;font-family:Sora,sans-serif" />
          </div>

          <div style="padding:16px;background:var(--grey1);border-radius:8px;margin-bottom:16px;font-size:12px;color:var(--grey3);line-height:1.6">
            <strong style="color:var(--navy)">Expected Format:</strong>
            <ul style="margin:8px 0 0 16px">
              <li>Sheet name: "Q4 Data"</li>
              <li>Columns: Company, Sector, Entry Year, Entry EV ($MM), Ownership %, 2024-Q4 Revenue ($MM), 2024-Q4 EBITDA ($MM), 2024-Q4 EBITDA Margin</li>
            </ul>
          </div>

          <button onclick="uploadFile()" style="background:var(--blue);color:white;border:none;padding:10px 20px;border-radius:8px;font-weight:600;cursor:pointer;width:100%;font-family:Sora,sans-serif;font-size:14px">
            <span id="uploadBtnText">Upload to Database</span>
          </button>
        </div>

        <div style="background:var(--grey1);border-radius:12px;padding:20px;display:none" id="uploadResultsDiv">
          <div style="font-weight:600;color:var(--navy);margin-bottom:16px">Upload Results</div>
          <div style="font-size:13px;line-height:1.8" id="uploadResults"></div>
        </div>
      </div>
    </div>

  </div>
</div>

<script>
// ── State ─────────────────────────────────────────────────────────────────────
let isAgentRunning = false;
let currentThinkingEl = null;

// ── Init ──────────────────────────────────────────────────────────────────────
document.addEventListener("DOMContentLoaded", () => {
  loadModel();
  loadSidebar();
  loadDashboard();
});

function loadModel() {
  fetch("/api/model").then(r=>r.json()).then(d=>{
    document.getElementById("modelBadge").textContent = d.model;
  });
}

function loadSidebar() {
  fetch("/api/portfolio").then(r=>r.json()).then(d=>{
    const list = document.getElementById("companyList");
    if (!d.rows || d.rows.length === 0) {
      list.innerHTML = '<div style="padding:16px;font-size:12px;color:var(--grey3)">No data</div>';
      return;
    }
    const SECTOR_COLORS = {
      "Healthcare":"#E74C3C","Construction":"#E67E22","Technology":"#3498DB",
      "Retail":"#9B59B6","Industrial":"#2ECC71","Consumer":"#F39C12",
      "Logistics":"#1ABC9C","Energy":"#F1C40F"
    };
    list.innerHTML = d.rows.map(row => {
      const [name, sector, rev, ebitda, margin] = row;
      const color = SECTOR_COLORS[sector] || "#95A5A6";
      const pct = margin ? (margin*100).toFixed(1)+"%" : "—";
      const short = name.replace(" Inc.","").replace(" Holdings","")
                        .replace(" Services","").replace(" Manufacturing","")
                        .replace(" Platform","");
      return `<div class="sidebar-item" onclick="askAbout('${name}')">
        <div class="sector-dot" style="background:${color}"></div>
        <span style="flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">${short}</span>
        <span class="margin-pill">${pct}</span>
      </div>`;
    }).join("");
  });
}

let portfolioData = [];

function loadDashboard() {
  // Load portfolio stats
  fetch("/api/portfolio/stats").then(r=>r.json()).then(stats=>{
    if (!stats.success) return;
    const statElements = document.querySelectorAll(".stat-card");
    statElements[0].innerHTML = `<div class="stat-label">Total Revenue</div><div class="stat-value">$${stats.total_revenue_mm.toFixed(0)}M</div><div class="stat-sub">2024-Q4</div>`;
    statElements[1].innerHTML = `<div class="stat-label">Avg EBITDA Margin</div><div class="stat-value">${(stats.avg_ebitda_margin*100).toFixed(1)}%</div><div class="stat-sub">of revenue</div>`;
    statElements[2].innerHTML = `<div class="stat-label"># Companies</div><div class="stat-value">${stats.company_count}</div>`;
    statElements[3].innerHTML = `<div class="stat-label">Top Performer</div><div class="stat-value" style="font-size:18px">${stats.top_performer_name || "—"}</div><div class="stat-sub">${stats.top_performer_margin ? (stats.top_performer_margin*100).toFixed(1)+"%" : "—"}</div>`;
  });

  // Load portfolio data for cards
  fetch("/api/portfolio").then(r=>r.json()).then(d=>{
    if (!d.rows || d.rows.length === 0) {
      document.getElementById("companyCards").innerHTML = '<div style="padding:40px;color:var(--grey3);text-align:center;grid-column:1/-1">No portfolio data</div>';
      return;
    }

    portfolioData = [];
    d.rows.forEach(row => {
      portfolioData.push({
        id: row[0],
        name: row[1],
        sector: row[2],
        revenue: parseFloat(row[3]),
        ebitda: parseFloat(row[4]),
        margin: parseFloat(row[5]),
        growth: row[6] ? parseFloat(row[6]) : 0
      });
    });

    sortAndRenderCards("margin-desc");
  });

  // Load comps data
  fetch("/api/comps").then(r=>r.json()).then(d=>{
    const wrap = document.getElementById("compsTableWrap");
    if (!d.rows || d.rows.length === 0) {
      wrap.innerHTML = '<div style="padding:40px;color:var(--grey3);text-align:center">No comps data yet</div>';
      return;
    }
    const cols = d.columns;
    let html = `<table class="data-table"><thead><tr>${cols.map(c=>`<th>${c.toUpperCase().replace("_"," ")}</th>`).join("")}</tr></thead><tbody>`;
    d.rows.slice(0,20).forEach(row => {
      html += "<tr>" + row.map((v,i) => {
        const col = cols[i];
        if (col === "sector_key") return `<td><span class="sector-badge">${v||"—"}</span></td>`;
        if (col === "last_updated") return `<td style="font-size:10px;color:var(--grey3)">${v ? v.substring(0,10) : "—"}</td>`;
        if (col.includes("margin") || col.includes("growth")) {
          const n = parseFloat(v);
          const cls = n >= 0 ? "positive" : "negative";
          return `<td class="num ${cls}">${isNaN(n) ? v : (n*100).toFixed(1)+"%"}</td>`;
        }
        if (col.includes("ev_")) return `<td class="num">${parseFloat(v).toFixed(1)}x</td>`;
        if (col.includes("mm")) return `<td class="num">$${parseFloat(v).toFixed(0)}M</td>`;
        return `<td>${v === null ? "—" : v}</td>`;
      }).join("") + "</tr>";
    });
    html += "</tbody></table>";
    if (d.rows.length > 20) html += `<div class="data-table-count">Showing first 20 of ${d.rows.length} comps</div>`;
    wrap.innerHTML = html;
  });
}

function sortAndRenderCards(sortBy) {
  let sorted = [...portfolioData];

  if (sortBy === "margin-desc") sorted.sort((a,b) => (b.margin||0) - (a.margin||0));
  else if (sortBy === "revenue-desc") sorted.sort((a,b) => (b.revenue||0) - (a.revenue||0));
  else if (sortBy === "growth-desc") sorted.sort((a,b) => (b.growth||0) - (a.growth||0));
  else if (sortBy === "name") sorted.sort((a,b) => a.name.localeCompare(b.name));

  const cardsHtml = sorted.map(co => {
    const marginCls = (co.margin||0) >= 0.20 ? "positive" : "negative";
    const growthCls = (co.growth||0) >= 0.05 ? "positive" : "negative";
    return `
      <div class="company-card" onclick="openCompanyModal(${co.id})">
        <div class="company-card-header">
          <div>
            <div class="company-name">${co.name}</div>
            <span class="sector-badge">${co.sector}</span>
          </div>
        </div>
        <div class="company-metrics">
          <div class="metric-item">
            <div class="metric-item-label">Revenue</div>
            <div class="metric-item-value">$${co.revenue.toFixed(0)}M</div>
          </div>
          <div class="metric-item">
            <div class="metric-item-label">EBITDA Margin</div>
            <div class="metric-item-value ${marginCls}">${(co.margin*100).toFixed(1)}%</div>
          </div>
          <div class="metric-item">
            <div class="metric-item-label">EBITDA</div>
            <div class="metric-item-value">$${(co.revenue*co.margin).toFixed(0)}M</div>
          </div>
          <div class="metric-item">
            <div class="metric-item-label">YoY Growth</div>
            <div class="metric-item-value ${growthCls}">${(co.growth*100).toFixed(1)}%</div>
          </div>
        </div>
      </div>
    `;
  }).join("");

  document.getElementById("companyCards").innerHTML = cardsHtml || '<div style="padding:40px;color:var(--grey3);text-align:center">No companies to display</div>';
}

function openCompanyModal(companyId) {
  const company = portfolioData.find(c => c.id === companyId);
  if (!company) return;

  document.getElementById("modalCompanyName").textContent = company.name;
  document.getElementById("modalSector").textContent = company.sector;

  // Load full company details
  fetch(`/api/company/${companyId}`)
    .then(r => r.json())
    .then(detail => {
      if (!detail.success) {
        console.log("Note: Company detail endpoint expects numeric ID, will show summary instead");
        document.getElementById("modal-revenue").textContent = `$${company.revenue.toFixed(0)}M`;
        document.getElementById("modal-ebitda").textContent = `$${(company.revenue*company.margin).toFixed(0)}M`;
        document.getElementById("modal-margin").textContent = `${(company.margin*100).toFixed(1)}%`;
        document.getElementById("modal-growth").textContent = `${(company.growth*100).toFixed(1)}%`;
        document.getElementById("modal-entry-ev").textContent = "—";
        document.getElementById("modal-ownership").textContent = "—";
        document.getElementById("kpisSection").style.display = "none";
        document.getElementById("compsSection").style.display = "none";
      } else {
        const co = detail.company;
        const fin = detail.financials.find(f => f.period === "2024-Q4") || {};
        document.getElementById("modal-revenue").textContent = fin.revenue_mm ? `$${fin.revenue_mm.toFixed(0)}M` : "—";
        document.getElementById("modal-ebitda").textContent = fin.ebitda_mm ? `$${fin.ebitda_mm.toFixed(0)}M` : "—";
        document.getElementById("modal-margin").textContent = fin.ebitda_margin ? `${(fin.ebitda_margin*100).toFixed(1)}%` : "—";
        document.getElementById("modal-growth").textContent = fin.yoy_growth ? `${(fin.yoy_growth*100).toFixed(1)}%` : "—";
        document.getElementById("modal-entry-ev").textContent = co.entry_ev_mm ? `$${co.entry_ev_mm.toFixed(0)}M` : "—";
        document.getElementById("modal-ownership").textContent = co.ownership_pct ? `${(co.ownership_pct*100).toFixed(0)}%` : "—";

        if (detail.kpis && detail.kpis.length > 0) {
          const kpi2024 = detail.kpis.filter(k => k.period === "2024-Q4");
          if (kpi2024.length > 0) {
            let kpiHtml = "";
            kpi2024.forEach(k => {
              kpiHtml += `<tr><td style="font-weight:500">${k.kpi_name}</td><td>${k.kpi_value}</td><td style="color:var(--grey3);font-size:12px">${k.kpi_unit}</td></tr>`;
            });
            document.getElementById("kpisTableBody").innerHTML = kpiHtml;
            document.getElementById("kpisSection").style.display = "block";
          }
        } else {
          document.getElementById("kpisSection").style.display = "none";
        }

        if (detail.comps && detail.comps.length > 0) {
          document.getElementById("compsSection").style.display = "block";
        } else {
          document.getElementById("compsSection").style.display = "none";
        }

        renderTrendCharts(detail.financials);
      }
    })
    .catch(err => {
      console.log("Could not load company detail:", err);
      document.getElementById("modal-revenue").textContent = `$${company.revenue.toFixed(0)}M`;
      document.getElementById("modal-ebitda").textContent = `$${(company.revenue*company.margin).toFixed(0)}M`;
      document.getElementById("modal-margin").textContent = `${(company.margin*100).toFixed(1)}%`;
      document.getElementById("modal-growth").textContent = `${(company.growth*100).toFixed(1)}%`;
    });

  document.getElementById("companyModal").classList.add("show");
}

function closeModal() {
  document.getElementById("companyModal").classList.remove("show");
}

function renderTrendCharts(financials) {
  const sorted = financials.sort((a,b) => a.period.localeCompare(b.period));
  const labels = sorted.map(f => f.period);
  const revenues = sorted.map(f => f.revenue_mm || 0);
  const ebitdas = sorted.map(f => f.ebitda_mm || 0);
  const margins = sorted.map(f => (f.ebitda_margin || 0) * 100);

  setTimeout(() => {
    const ctx1 = document.getElementById("trendChart")?.getContext("2d");
    if (ctx1) {
      new Chart(ctx1, {
        type: "line",
        data: {
          labels,
          datasets: [
            { label: "Revenue ($MM)", data: revenues, borderColor: "var(--blue)", backgroundColor: "rgba(30,111,217,0.1)", tension: 0.4, borderWidth: 2 },
            { label: "EBITDA ($MM)", data: ebitdas, borderColor: "var(--teal)", backgroundColor: "rgba(10,191,188,0.1)", tension: 0.4, borderWidth: 2 }
          ]
        },
        options: { responsive: true, maintainAspectRatio: false, plugins: { legend: { position: "top" } }, scales: { y: { beginAtZero: true } } }
      });
    }

    const ctx2 = document.getElementById("marginChart")?.getContext("2d");
    if (ctx2) {
      new Chart(ctx2, {
        type: "line",
        data: {
          labels,
          datasets: [{ label: "EBITDA Margin (%)", data: margins, borderColor: "var(--green)", backgroundColor: "rgba(39,174,96,0.1)", tension: 0.4, borderWidth: 2, fill: true }]
        },
        options: { responsive: true, maintainAspectRatio: false, plugins: { legend: { position: "top" } }, scales: { y: { beginAtZero: true, max: 100 } } }
      });
    }
  }, 100);
}

function loadCompsTable() {
  // Populate company selector dropdown
  const selector = document.getElementById("companySelector");
  selector.innerHTML = '<option value="">-- Choose a company --</option>';

  portfolioData.forEach(co => {
    const opt = document.createElement("option");
    opt.value = co.id;
    opt.textContent = `${co.name} (${co.sector})`;
    selector.appendChild(opt);
  });
}

function selectCompanyForComparison(companyId) {
  if (!companyId) {
    document.getElementById("selectedCompanyCard").style.display = "none";
    document.getElementById("comparablesSection").style.display = "none";
    document.getElementById("compsEmptyState").style.display = "block";
    return;
  }

  const company = portfolioData.find(c => c.id == companyId);
  if (!company) return;

  // Show company card
  document.getElementById("selectedCompanyName").textContent = company.name;
  document.getElementById("selectedCompanySector").textContent = company.sector;
  document.getElementById("selectedRevenue").textContent = `$${company.revenue.toFixed(0)}M`;
  document.getElementById("selectedMargin").textContent = `${(company.margin*100).toFixed(1)}%`;
  document.getElementById("selectedGrowth").textContent = `${(company.growth*100).toFixed(1)}%`;
  document.getElementById("selectedEV").textContent = company.entry_ev_mm ? `$${company.entry_ev_mm.toFixed(0)}M` : "—";
  document.getElementById("selectedCompanyCard").style.display = "block";
  document.getElementById("compsEmptyState").style.display = "none";

  // Load sector comparables
  fetch(`/api/comps`).then(r=>r.json()).then(d=>{
    if (!d.rows || d.rows.length === 0) {
      document.getElementById("comparablesSection").style.display = "none";
      return;
    }

    const cols = d.columns;
    const comps = [];
    d.rows.forEach(row => {
      const comp = {};
      row.forEach((val, idx) => {
        comp[cols[idx]] = val;
      });
      comps.push(comp);
    });

    // Filter to same sector (case-insensitive)
    const companySectorLower = (company.sector || "").toLowerCase();
    const sectorComps = comps.filter(c => (c.sector_key || "").toLowerCase() === companySectorLower);
    renderComparisonView(company, sectorComps);
  });
}

function renderComparisonView(company, comps) {
  const tbody = document.getElementById("comparablesTableBody");
  tbody.innerHTML = "";

  if (!comps || comps.length === 0) {
    document.getElementById("comparablesSection").style.display = "none";
    return;
  }

  document.getElementById("comparablesSection").style.display = "block";

  // First, add our company as the baseline
  const ourRow = document.createElement("tr");
  ourRow.style.background = "rgba(30,111,217,0.15)";
  ourRow.style.fontWeight = "600";
  ourRow.style.borderBottom = "2px solid var(--blue)";

  // Company name (ours)
  const ourNameCell = document.createElement("td");
  ourNameCell.innerHTML = `<strong>${company.name}</strong><br><span style="color:var(--blue);font-size:11px;font-weight:600">YOUR COMPANY</span>`;
  ourRow.appendChild(ourNameCell);

  // Ticker (N/A for ours)
  const ourTickerCell = document.createElement("td");
  ourTickerCell.textContent = "—";
  ourTickerCell.style.textAlign = "right";
  ourRow.appendChild(ourTickerCell);

  // Revenue (ours)
  const ourRevCell = document.createElement("td");
  ourRevCell.textContent = `$${company.revenue.toFixed(0)}M`;
  ourRevCell.style.textAlign = "right";
  ourRevCell.style.color = "var(--navy)";
  ourRow.appendChild(ourRevCell);

  // Growth (ours)
  const ourGrowthCell = document.createElement("td");
  ourGrowthCell.textContent = `${(company.growth*100).toFixed(1)}%`;
  ourGrowthCell.style.textAlign = "right";
  ourGrowthCell.style.color = "var(--navy)";
  ourRow.appendChild(ourGrowthCell);

  // Margin (ours)
  const ourMarginCell = document.createElement("td");
  ourMarginCell.textContent = `${(company.margin*100).toFixed(1)}%`;
  ourMarginCell.style.textAlign = "right";
  ourMarginCell.style.color = "var(--navy)";
  ourRow.appendChild(ourMarginCell);

  // EV/EBITDA (N/A for ours)
  const ourEVCell = document.createElement("td");
  ourEVCell.textContent = "—";
  ourEVCell.style.textAlign = "right";
  ourRow.appendChild(ourEVCell);

  // EV/Revenue (N/A for ours)
  const ourEVRevCell = document.createElement("td");
  ourEVRevCell.textContent = "—";
  ourEVRevCell.style.textAlign = "right";
  ourRow.appendChild(ourEVRevCell);

  tbody.appendChild(ourRow);

  // Now add sector comparables
  comps.forEach(comp => {
    const row = document.createElement("tr");

    // Company name
    const nameCell = document.createElement("td");
    nameCell.innerHTML = `<strong>${comp.name}</strong><br><span style="color:var(--grey3);font-size:11px">${comp.sector_key}</span>`;
    row.appendChild(nameCell);

    // Ticker
    const tickerCell = document.createElement("td");
    tickerCell.textContent = comp.ticker || "—";
    tickerCell.style.textAlign = "right";
    row.appendChild(tickerCell);

    // Revenue comparison
    const revenue = parseFloat(comp.market_cap_mm) || 0;
    const revCell = document.createElement("td");
    revCell.textContent = revenue > 0 ? `$${revenue.toFixed(0)}M` : "—";
    revCell.style.textAlign = "right";
    if (revenue > 0) {
      revCell.style.color = revenue > company.revenue ? "var(--green)" : "var(--red)";
      revCell.style.fontWeight = "600";
    }
    row.appendChild(revCell);

    // Growth comparison
    const growth = parseFloat(comp.revenue_growth) || 0;
    const growthCell = document.createElement("td");
    growthCell.textContent = !isNaN(growth) && growth !== null ? `${(growth*100).toFixed(1)}%` : "—";
    growthCell.style.textAlign = "right";
    if (!isNaN(growth) && growth !== null) {
      growthCell.style.color = growth > company.growth ? "var(--green)" : "var(--red)";
      growthCell.style.fontWeight = "600";
    } else {
      growthCell.style.color = "var(--grey3)";
    }
    row.appendChild(growthCell);

    // EBITDA Margin comparison
    const margin = parseFloat(comp.ebitda_margin) || 0;
    const marginCell = document.createElement("td");
    marginCell.textContent = !isNaN(margin) && margin !== null ? `${(margin*100).toFixed(1)}%` : "—";
    marginCell.style.textAlign = "right";
    if (!isNaN(margin) && margin !== null) {
      marginCell.style.color = margin > company.margin ? "var(--green)" : "var(--red)";
      marginCell.style.fontWeight = "600";
    } else {
      marginCell.style.color = "var(--grey3)";
    }
    row.appendChild(marginCell);

    // EV/EBITDA
    const evEbitda = parseFloat(comp.ev_ebitda) || null;
    const evEbitdaCell = document.createElement("td");
    evEbitdaCell.textContent = evEbitda ? `${evEbitda.toFixed(1)}x` : "—";
    evEbitdaCell.style.textAlign = "right";
    evEbitdaCell.style.color = evEbitda ? "var(--teal)" : "var(--grey3)";
    row.appendChild(evEbitdaCell);

    // EV/Revenue
    const evRevenue = parseFloat(comp.ev_revenue) || null;
    const evRevenueCell = document.createElement("td");
    evRevenueCell.textContent = evRevenue ? `${evRevenue.toFixed(1)}x` : "—";
    evRevenueCell.style.textAlign = "right";
    evRevenueCell.style.color = evRevenue ? "var(--teal)" : "var(--grey3)";
    row.appendChild(evRevenueCell);

    tbody.appendChild(row);
  });
}

// ── Tab switching ─────────────────────────────────────────────────────────────
function switchTab(tab) {
  document.querySelectorAll(".panel").forEach(p => p.classList.remove("active"));
  document.querySelectorAll(".tab-btn").forEach(b => b.classList.remove("active"));
  document.getElementById(`panel-${tab}`).classList.add("active");
  event.target.classList.add("active");
  if (tab === "comps") loadCompsTable();
}

// ── Chat ──────────────────────────────────────────────────────────────────────
function sendSuggestion(el) {
  document.getElementById("chatInput").value = el.textContent;
  sendMessage();
}

function askAbout(name) {
  document.getElementById("chatInput").value = `Give me a full performance summary for ${name}`;
  switchTab("chat");
  document.querySelectorAll(".tab-btn")[0].classList.add("active");
  document.querySelectorAll(".tab-btn")[1].classList.remove("active");
  sendMessage();
}

function sendMessage() {
  const input = document.getElementById("chatInput");
  const question = input.value.trim();
  if (!question || isAgentRunning) return;

  // Clear empty state
  const emptyEl = document.querySelector(".chat-empty");
  if (emptyEl) emptyEl.remove();

  appendMsg("user", question);
  input.value = "";
  input.style.height = "auto";

  const thinkingEl = appendThinking();
  currentThinkingEl = thinkingEl;
  isAgentRunning = true;
  document.getElementById("sendBtn").disabled = true;

  const es = new EventSource(`/api/ask?q=${encodeURIComponent(question)}`);
  let toolTraceEl = null;

  es.onmessage = (e) => {
    const data = JSON.parse(e.data);

    if (data.type === "tool_start") {
      if (thinkingEl && thinkingEl.parentNode) thinkingEl.remove();
      toolTraceEl = appendToolTrace(data.tool, data.detail, true);
    }
    else if (data.type === "tool_done") {
      if (toolTraceEl) updateToolTrace(toolTraceEl, data.rows);
      toolTraceEl = null;
    }
    else if (data.type === "table") {
      appendDataTable(data.columns, data.rows, data.count);
    }
    else if (data.type === "answer") {
      if (thinkingEl && thinkingEl.parentNode) thinkingEl.remove();
      appendMsg("agent", data.msg);
      es.close();
      isAgentRunning = false;
      document.getElementById("sendBtn").disabled = false;
    }
    else if (data.type === "error") {
      if (thinkingEl && thinkingEl.parentNode) thinkingEl.remove();
      appendMsg("agent", `**Error:** ${data.msg}`);
      es.close();
      isAgentRunning = false;
      document.getElementById("sendBtn").disabled = false;
    }
  };

  es.onerror = () => {
    if (thinkingEl && thinkingEl.parentNode) thinkingEl.remove();
    es.close();
    isAgentRunning = false;
    document.getElementById("sendBtn").disabled = false;
  };
}

function appendMsg(role, text) {
  const msgs = document.getElementById("chatMessages");
  const div = document.createElement("div");
  div.className = `msg ${role}`;
  if (role === "agent") {
    div.innerHTML = `<div class="msg-label">Agent</div>
      <div class="msg-bubble">${marked.parse(text)}</div>`;
  } else {
    div.innerHTML = `<div class="msg-label">You</div>
      <div class="msg-bubble">${escHtml(text)}</div>`;
  }
  msgs.appendChild(div);
  msgs.scrollTop = msgs.scrollHeight;
  return div;
}

function appendThinking() {
  const msgs = document.getElementById("chatMessages");
  const div = document.createElement("div");
  div.className = "msg agent";
  div.innerHTML = `<div class="msg-label">Agent</div>
    <div class="thinking"><div class="dot"></div><div class="dot"></div><div class="dot"></div></div>`;
  msgs.appendChild(div);
  msgs.scrollTop = msgs.scrollHeight;
  return div;
}

function appendToolTrace(toolName, detail, running) {
  const msgs = document.getElementById("chatMessages");
  const div = document.createElement("div");
  div.className = "tool-trace";
  div.innerHTML = `<div class="tool-trace-row">
    ${running ? '<div class="tool-spinner"></div>' : '<span class="tool-done-icon">✓</span>'}
    <span class="tool-name">${toolName}</span>
    <span class="tool-detail">${escHtml(detail||"")}</span>
  </div>`;
  msgs.appendChild(div);
  msgs.scrollTop = msgs.scrollHeight;
  return div;
}

function updateToolTrace(el, rows) {
  const row = el.querySelector(".tool-trace-row");
  if (!row) return;
  const spinner = row.querySelector(".tool-spinner");
  if (spinner) { const done = document.createElement("span"); done.className="tool-done-icon"; done.textContent="✓"; spinner.replaceWith(done); }
  if (rows) { const detail = row.querySelector(".tool-detail"); if (detail) detail.textContent += ` → ${rows} rows`; }
}

function appendDataTable(columns, rows, count) {
  const msgs = document.getElementById("chatMessages");
  const wrap = document.createElement("div");
  wrap.className = "data-table-wrap";
  let html = `<table class="data-table"><thead><tr>${columns.map(c=>`<th>${c}</th>`).join("")}</tr></thead><tbody>`;
  rows.forEach(row => {
    html += "<tr>" + row.map(v => `<td>${v === null ? "—" : v}</td>`).join("") + "</tr>";
  });
  html += "</tbody></table>";
  if (count > rows.length) html += `<div class="data-table-count">Showing ${rows.length} of ${count} rows</div>`;
  wrap.innerHTML = html;
  msgs.appendChild(wrap);
  msgs.scrollTop = msgs.scrollHeight;
}

// ── Comps fetch ───────────────────────────────────────────────────────────────
function triggerFetchComps() {
  const log = document.getElementById("fetchLog");
  log.innerHTML = '<div style="color:var(--grey3);font-size:12px">Fetching market data from Yahoo Finance...</div>';

  const es = new EventSource("/api/fetch_comps");
  es.onmessage = (e) => {
    const data = JSON.parse(e.data);
    if (data.done) {
      es.close();
      log.innerHTML += `<div style="color:var(--green);margin-top:8px">✓ Complete - Market data updated</div>`;
      setTimeout(() => {
        loadCompsTable();
        loadSidebar();
        loadDashboard();
      }, 500);
      return;
    }
    log.innerHTML += `<div style="font-size:12px">${escHtml(data.msg)}</div>`;
    log.scrollTop = log.scrollHeight;
  };
  es.onerror = () => {
    es.close();
    log.innerHTML += `<div style="color:var(--red);margin-top:8px">Error fetching data</div>`;
  };
}

function toggleFetchPanel() {
  const log = document.getElementById("fetchLog");
  const showing = log.style.display !== "none";
  if (showing) {
    log.style.display = "none";
  } else {
    log.style.display = "block";
    log.innerHTML = "";
    triggerFetchComps();
  }
}

function uploadFile() {
  const fileInput = document.getElementById("uploadFile");
  if (!fileInput.files.length) {
    alert("Please select a file");
    return;
  }

  const file = fileInput.files[0];
  if (!file.name.endsWith(".xlsx")) {
    alert("Please select an .xlsx file");
    return;
  }

  const formData = new FormData();
  formData.append("file", file);

  const btn = document.querySelector("#panel-upload button");
  const btnText = document.getElementById("uploadBtnText");
  btn.disabled = true;
  btnText.textContent = "Uploading...";

  fetch("/api/upload", { method: "POST", body: formData })
    .then(r => r.json())
    .then(data => {
      btn.disabled = false;
      btnText.textContent = "Upload to Database";

      const resultsDiv = document.getElementById("uploadResultsDiv");
      const results = document.getElementById("uploadResults");

      if (!data.success) {
        results.innerHTML = `<div style="color:var(--red)">❌ Error: ${escHtml(data.error)}</div>`;
      } else {
        let html = `<div style="color:var(--green)">✓ Upload complete: ${data.uploaded} of ${data.total} companies</div><br>`;
        (data.results || []).forEach(r => {
          if (r.success) {
            html += `<div style="color:var(--green)">✓ ${escHtml(r.company)} (${r.action})</div>`;
          } else {
            html += `<div style="color:var(--red)">✗ ${escHtml(r.company)}: ${escHtml(r.error)}</div>`;
          }
        });
        results.innerHTML = html;
        fileInput.value = "";
        loadDashboard();  // Refresh dashboard
      }
      resultsDiv.style.display = "block";
    })
    .catch(err => {
      btn.disabled = false;
      btnText.textContent = "Upload to Database";
      const resultsDiv = document.getElementById("uploadResultsDiv");
      document.getElementById("uploadResults").innerHTML = `<div style="color:var(--red)">❌ Error: ${escHtml(err.message)}</div>`;
      resultsDiv.style.display = "block";
    });
}

function escHtml(s) {
  return String(s||"").replace(/&/g,"&amp;").replace(/</g,"&lt;").replace(/>/g,"&gt;");
}
</script>
</body>
</html>'''


# ── Routes ─────────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template_string(HTML)

@app.route("/api/model")
def api_model():
    return jsonify({"model": MODEL})

@app.route("/api/health")
def api_health():
    env_keys = ["ANTHROPIC_API_KEY", "OPENAI_API_KEY", "GOOGLE_API_KEY"]
    env_status = {key: bool(os.getenv(key)) for key in env_keys}
    return jsonify({
        "model": MODEL,
        "env_status": env_status,
        "tool_count": len(TOOL_DEFINITIONS),
    })

@app.route("/api/portfolio")
def api_portfolio():
    return jsonify(get_portfolio_summary())

@app.route("/api/comp_count")
def api_comp_count():
    return jsonify({"count": get_comp_count()})

@app.route("/api/comps")
def api_comps():
    return jsonify(get_comps_summary())

@app.route("/api/price_history/<ticker>")
def api_price_history(ticker):
    return jsonify(get_price_history(ticker.upper()))

@app.route("/api/ask")
def api_ask():
    question = request.args.get("q", "")
    if not question:
        return Response("data: {}\n\n", mimetype="text/event-stream")

    q = queue.Queue()

    def run():
        try:
            run_agent_streaming(question, q)
        except Exception as e:
            q.put({"type": "error", "msg": str(e)})

    threading.Thread(target=run, daemon=True).start()

    def generate():
        while True:
            try:
                item = q.get(timeout=120)
                yield f"data: {json.dumps(item)}\n\n"
                if item.get("type") in ("answer", "error"):
                    break
            except queue.Empty:
                yield f"data: {json.dumps({'type':'error','msg':'Timeout'})}\n\n"
                break

    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control":"no-cache","X-Accel-Buffering":"no"})

@app.route("/api/fetch_comps")
def api_fetch_comps():
    from fetch_comps import fetch_all_comps, ALL_TICKERS
    log_q = queue.Queue()

    def run():
        import sys
        class QueueLogger:
            def write(self, msg):
                if msg.strip():
                    log_q.put({"msg": msg.strip()})
            def flush(self): pass

        old_stdout = sys.stdout
        sys.stdout = QueueLogger()
        try:
            fetch_all_comps(refresh=True)
        finally:
            sys.stdout = old_stdout
        log_q.put({"done": True})

    threading.Thread(target=run, daemon=True).start()

    def generate():
        while True:
            try:
                item = log_q.get(timeout=120)
                yield f"data: {json.dumps(item)}\n\n"
                if item.get("done"):
                    break
            except queue.Empty:
                yield f"data: {json.dumps({'done':True})}\n\n"
                break

    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control":"no-cache","X-Accel-Buffering":"no"})


@app.route("/api/company/<int:company_id>")
def api_company_detail(company_id):
    result = get_company_detail(company_id)
    if not result.get("success"):
        return jsonify({"success": False, "error": result.get("error", "Company not found")}), 404

    company = result["company"]
    financials = result["financials"]
    kpis = result["kpis"]

    sector = company.get("sector", "")
    comps = get_company_comps(sector)
    comps_list = comps.get("comps", []) if comps.get("success") else []

    return jsonify({
        "success": True,
        "company": company,
        "financials": financials,
        "kpis": kpis,
        "comps": comps_list
    })


@app.route("/api/portfolio/stats")
def api_portfolio_stats():
    q = execute_query("""
        SELECT
            COUNT(DISTINCT c.id) as company_count,
            SUM(CASE WHEN qf.period = '2024-Q4' THEN qf.revenue_mm ELSE 0 END) as total_revenue_mm,
            SUM(CASE WHEN qf.period = '2024-Q4' THEN qf.ebitda_mm ELSE 0 END) as total_ebitda_mm,
            AVG(CASE WHEN qf.period = '2024-Q4' THEN qf.ebitda_margin ELSE NULL END) as avg_ebitda_margin,
            MAX(CASE WHEN qf.period = '2024-Q4' THEN qf.ebitda_margin ELSE NULL END) as max_ebitda_margin
        FROM companies c
        LEFT JOIN quarterly_financials qf ON c.id = qf.company_id
    """)

    if not q.get("success"):
        return jsonify({"success": False, "error": q.get("error")}), 500

    stats = q["rows"][0] if q["rows"] else [0, 0, 0, 0, 0]
    company_count = int(stats[0]) if stats[0] else 0
    total_revenue = round(float(stats[1]) if stats[1] else 0, 2)
    total_ebitda = round(float(stats[2]) if stats[2] else 0, 2)
    avg_margin = round(float(stats[3]) if stats[3] else 0, 4)

    top_performer = execute_query("""
        SELECT c.name, qf.ebitda_margin
        FROM companies c
        JOIN quarterly_financials qf ON c.id = qf.company_id
        WHERE qf.period = '2024-Q4'
        ORDER BY qf.ebitda_margin DESC NULLS LAST
        LIMIT 1
    """)

    top_name = ""
    top_margin = 0
    if top_performer.get("success") and top_performer.get("rows"):
        top_name = top_performer["rows"][0][0]
        top_margin = round(float(top_performer["rows"][0][1]), 4) if top_performer["rows"][0][1] else 0

    return jsonify({
        "success": True,
        "company_count": company_count,
        "total_revenue_mm": total_revenue,
        "total_ebitda_mm": total_ebitda,
        "avg_ebitda_margin": avg_margin,
        "top_performer_name": top_name,
        "top_performer_margin": top_margin
    })


@app.route("/api/upload", methods=["POST"])
def api_upload():
    if "file" not in request.files:
        return jsonify({"success": False, "error": "No file provided"}), 400

    file = request.files["file"]
    if file.filename == "":
        return jsonify({"success": False, "error": "No file selected"}), 400

    if not file.filename.endswith(".xlsx"):
        return jsonify({"success": False, "error": "Only .xlsx files are supported"}), 400

    try:
        import openpyxl
        from io import BytesIO

        file_bytes = BytesIO(file.read())
        wb = openpyxl.load_workbook(file_bytes)

        if "Q4 Data" not in wb.sheetnames:
            return jsonify({"success": False, "error": "Sheet 'Q4 Data' not found in workbook"}), 400

        ws = wb["Q4 Data"]

        results = []
        headers = {}

        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value:
                headers[cell.value.lower()] = col_idx

        required_fields = ["company", "sector", "entry year", "entry ev ($mm)", "ownership %", "2024-q4 revenue ($mm)", "2024-q4 ebitda ($mm)", "2024-q4 ebitda margin"]
        missing = [f for f in required_fields if f.lower() not in headers]

        if missing:
            return jsonify({"success": False, "error": f"Missing columns: {', '.join(missing)}"}), 400

        for row_idx in range(2, ws.max_row + 1):
            try:
                company_name = ws.cell(row_idx, headers.get("company", 1)).value
                if not company_name:
                    continue

                sector = ws.cell(row_idx, headers.get("sector", 1)).value or "Unknown"
                entry_year = ws.cell(row_idx, headers.get("entry year", 1)).value
                entry_ev = ws.cell(row_idx, headers.get("entry ev ($mm)", 1)).value
                ownership = ws.cell(row_idx, headers.get("ownership %", 1)).value
                revenue = ws.cell(row_idx, headers.get("2024-q4 revenue ($mm)", 1)).value
                ebitda = ws.cell(row_idx, headers.get("2024-q4 ebitda ($mm)", 1)).value
                ebitda_margin = ws.cell(row_idx, headers.get("2024-q4 ebitda margin", 1)).value

                entry_year = int(entry_year) if entry_year else 2020
                entry_ev = float(entry_ev) if entry_ev else 0
                ownership = float(ownership) if ownership else 0.5
                revenue = float(revenue) if revenue else 0
                ebitda = float(ebitda) if ebitda else 0
                ebitda_margin = float(ebitda_margin) if ebitda_margin else 0

                company_result = upsert_company(company_name, sector, entry_year, entry_ev, ownership)
                if not company_result.get("success"):
                    results.append({"company": company_name, "success": False, "error": company_result.get("error")})
                    continue

                company_id = company_result["company_id"]

                fin_result = upsert_financials(company_id, "2024-Q4", revenue, ebitda, ebitda_margin)
                if not fin_result.get("success"):
                    results.append({"company": company_name, "success": False, "error": f"Financials: {fin_result.get('error')}"})
                    continue

                results.append({
                    "company": company_name,
                    "success": True,
                    "company_id": company_id,
                    "action": company_result.get("action", "unknown")
                })

            except Exception as e:
                results.append({"company": company_name or f"Row {row_idx}", "success": False, "error": str(e)})

        successful = sum(1 for r in results if r.get("success"))
        return jsonify({
            "success": True,
            "uploaded": successful,
            "total": len(results),
            "results": results
        })

    except Exception as e:
        return jsonify({"success": False, "error": f"Upload failed: {str(e)}"}), 500


if __name__ == "__main__":
    seed_sample_data()
    print(f"\nPortfolio Intelligence Agent")
    print(f"Model: {MODEL}")
    print(f"Open: http://localhost:5000\n")
    app.run(debug=False, port=5000, threaded=True)
